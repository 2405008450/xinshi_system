"""笔译项目 Excel 导出。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Any, Callable, Iterable, Iterator, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from crud import get_translation_projects


EXPORT_BATCH_SIZE = 500
EXPORT_MAX_ROWS_PER_SHEET = 50_000

TIME_FIELD_LABELS = {
    "customer_reception_time": "客户接单时间",
    "customer_deadline_time": "客户交稿时间",
    "created_at": "创建时间",
}

STATUS_LABELS = {
    "pending": "待确认",
    "pending_confirmation": "待确认",
    "confirmed": "已确认",
    "in_progress": "已确认",
    "organized": "已整理",
    "translator_assigned": "已排译员",
    "sent_to_translator": "已发译员",
    "translator_returned": "译员发回",
    "special_checked": "已专检",
    "typeset": "已排版",
    "special_checked_typeset": "已专检排版",
    "reviewed": "已审核",
    "sent_to_client": "已发客户",
    "completed": "已发客户",
    "client_feedback": "客户反馈",
    "feedback_sent_to_client": "反馈后发客户",
    "cancelled": "已取消",
    "terminated": "已取消",
    "partially_cancelled": "已部分取消",
    "paused": "已暂停",
}

WORD_COUNT_DIMENSIONS = (
    ("company", "我司"),
    ("customer", "客户"),
    ("translator_estimate", "译员预估"),
)
WORD_COUNT_METRICS = (
    ("words", "字数"),
    ("characters_no_spaces", "字符数（不计空格）"),
    ("cjk_chars_korean_words", "中文字符和朝鲜语单词"),
    ("foreign_words", "外文字数"),
    ("documents", "份数"),
    ("pages", "页数"),
)

_ILLEGAL_EXCEL_TEXT = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")


class TranslationExportEmptyError(ValueError):
    """所选条件没有可导出的母订单。"""


class TranslationExportLimitError(ValueError):
    """某个工作表超过安全导出行数。"""


@dataclass(frozen=True)
class ExportColumn:
    label: str
    getter: Callable[[Any], Any]
    kind: str = "text"
    width: int = 18


def _read(value: Any, key: str, default: Any = None) -> Any:
    if isinstance(value, dict):
        return value.get(key, default)
    return getattr(value, key, default)


def _attr(name: str) -> Callable[[Any], Any]:
    return lambda item: _read(item, name)


def _safe_text(value: Any) -> str | None:
    if value is None:
        return None
    text = _ILLEGAL_EXCEL_TEXT.sub("", str(value))[:32767]
    if text.startswith(_FORMULA_PREFIXES):
        text = f"'{text}"
    return text


def _status_label(value: Any) -> str | None:
    if value in (None, ""):
        return None
    return STATUS_LABELS.get(str(value), str(value))


def _role_name(project: Any, role_code: str) -> str:
    for assignment in _read(project, "role_assignments", []) or []:
        if _read(assignment, "role_code") != role_code:
            continue
        return _read(assignment, "assignee_name") or "角色池"
    return "角色池"


def _assigned_translators(item: Any) -> str | None:
    assignments = _read(item, "assigned_translators", []) or []
    values = []
    for assignment in assignments:
        name = _read(assignment, "translator_name") or ""
        scope = _read(assignment, "translation_scope") or ""
        if name:
            values.append(f"{name}（{scope}）" if scope else name)
    if values:
        return "\n".join(values)
    return _read(item, "translator_name") or None


def _translator_return_times(item: Any) -> str | None:
    values = []
    for assignment in _read(item, "assigned_translators", []) or []:
        return_time = _read(assignment, "translator_return_time")
        if not return_time:
            continue
        name = _read(assignment, "translator_name") or "译员"
        if isinstance(return_time, (date, datetime)):
            display_time = return_time.strftime("%Y-%m-%d %H:%M")
        else:
            display_time = str(return_time)
        values.append(f"{name}：{display_time}")
    return "\n".join(values) or None


def _translator_completion_remarks(item: Any) -> str | None:
    values = []
    for assignment in _read(item, "assigned_translators", []) or []:
        remarks = _read(assignment, "completion_remarks")
        if remarks:
            name = _read(assignment, "translator_name") or "译员"
            values.append(f"{name}：{remarks}")
    return "\n".join(values) or None


def _word_count_value(item: Any, dimension: str, metric: str) -> Any:
    matrix = _read(item, "word_count_matrix", {}) or {}
    values = _read(matrix, dimension, {}) or {}
    return _read(values, metric)


def _word_count_columns() -> list[ExportColumn]:
    return [
        ExportColumn(
            f"{dimension_label}-{metric_label}",
            lambda item, dimension=dimension, metric=metric: _word_count_value(item, dimension, metric),
            kind="integer",
            width=18 if metric not in {"characters_no_spaces", "cjk_chars_korean_words"} else 24,
        )
        for dimension, dimension_label in WORD_COUNT_DIMENSIONS
        for metric, metric_label in WORD_COUNT_METRICS
    ]


PROJECT_COLUMNS = [
    ExportColumn("订单号", _attr("order_no"), "identifier", 18),
    ExportColumn("项目名称", _attr("project_name"), width=32),
    ExportColumn("邮件主题预览", _attr("email_subject_preview"), width=36),
    ExportColumn("服务内容", _attr("service_content"), width=24),
    ExportColumn("任务类型", _attr("task_type")),
    ExportColumn("来源咨询 ID", _attr("consultation_id"), "identifier", 38),
    ExportColumn("客户简称", _attr("client_short_name"), width=24),
    ExportColumn("客户编号", _attr("client_code"), "identifier", 18),
    ExportColumn("客户单号", _attr("customer_order_no"), "identifier", 20),
    ExportColumn("项目经理", _attr("project_manager_name")),
    ExportColumn("项目专员", lambda item: _role_name(item, "project_specialist")),
    ExportColumn("项目助理", lambda item: _role_name(item, "project_assistant")),
    ExportColumn("排版专员", lambda item: _role_name(item, "layout_specialist")),
    ExportColumn("客户经理", _attr("client_manager")),
    ExportColumn("客户经理联系方式", _attr("manager_contact"), width=24),
    ExportColumn("状态", lambda item: _status_label(_read(item, "project_status")), width=16),
    ExportColumn("文本类型", _attr("file_type_secondary")),
    ExportColumn("翻译文本领域一级", _attr("project_file_translation_domain_level1"), width=22),
    ExportColumn("翻译文本领域二级", _attr("project_file_translation_domain_level2"), width=22),
    ExportColumn("文件类型一级", _attr("project_file_type_level1")),
    ExportColumn("文件类型二级", _attr("project_file_type_level2")),
    ExportColumn("文件格式", _attr("project_file_format")),
    ExportColumn("文件属性一级", _attr("project_file_attribute_level1")),
    ExportColumn("文件属性二级", _attr("project_file_attribute_level2")),
    ExportColumn("文件属性三级", _attr("project_file_attribute_level3")),
    ExportColumn("文件难度", _attr("project_file_difficulty")),
    ExportColumn("合同类型", _attr("project_contract_type")),
    ExportColumn("合同状态", _attr("project_contract_status")),
    ExportColumn("需提供报价单", lambda item: "是" if _read(item, "quotation_required", False) else "否", width=16),
    ExportColumn("报价单状态", _attr("quotation_status")),
    ExportColumn("报价单路径", _attr("quotation_path"), width=38),
    ExportColumn("客户专业要求", _attr("customer_requirement_professional"), width=36),
    ExportColumn("客户特殊要求", _attr("customer_requirement_special"), width=36),
    ExportColumn("翻译方向", _attr("language_pair"), width=24),
    ExportColumn("优先级", _attr("priority"), width=14),
    *_word_count_columns(),
    ExportColumn("客户接单时间", _attr("customer_reception_time"), "datetime", 20),
    ExportColumn("客户交稿时间", _attr("customer_deadline_time"), "datetime", 20),
    ExportColumn("发客户时间", _attr("sent_to_client_time"), "datetime", 20),
    ExportColumn("PM确认人 ID", _attr("pm_confirmed_by"), "identifier", 38),
    ExportColumn("客户反馈", _attr("client_feedback"), width=36),
    ExportColumn("大项目经理确认", _attr("major_project_manager_confirmation"), width=24),
    ExportColumn("已分配译员", _assigned_translators, width=28),
    ExportColumn("译员回稿时间", _translator_return_times, width=28),
    ExportColumn("译员任务完成情况", _translator_completion_remarks, width=36),
    ExportColumn("译员分配时间", _attr("translator_assignment_time"), "datetime", 20),
    ExportColumn("译员交付进度", _attr("translator_delivery_progress"), "percent", 16),
    ExportColumn("审校前 QC", _attr("pre_review_qc_progress"), "percent", 14),
    ExportColumn("审校 1", _attr("review1_progress"), "percent", 12),
    ExportColumn("审校 2", _attr("review2_progress"), "percent", 12),
    ExportColumn("审校后 QC", _attr("post_review_qc_progress"), "percent", 14),
    ExportColumn("排版进度", _attr("layout_progress"), "percent", 14),
    ExportColumn("整合进度", _attr("consolidation_progress"), "percent", 14),
    ExportColumn("创建时间", _attr("created_at"), "datetime", 20),
    ExportColumn("更新时间", _attr("updated_at"), "datetime", 20),
]

SUB_ORDER_COLUMNS = [
    ExportColumn("母订单号", lambda pair: _read(pair[0], "order_no"), "identifier", 18),
    ExportColumn("母项目名称", lambda pair: _read(pair[0], "project_name"), width=32),
    ExportColumn("子订单号", lambda pair: _read(pair[1], "sub_order_no"), "identifier", 20),
    ExportColumn("子项目名称", lambda pair: _read(pair[1], "sub_project_name"), width=32),
    ExportColumn("状态", lambda pair: _status_label(_read(pair[1], "status")), width=16),
    ExportColumn("文本类型", lambda pair: _read(pair[1], "file_type_secondary")),
    ExportColumn("翻译方向", lambda pair: _read(pair[1], "language_pair"), width=24),
    ExportColumn("优先级", lambda pair: _read(pair[1], "priority"), width=14),
    *[
        ExportColumn(
            column.label,
            lambda pair, getter=column.getter: getter(pair[1]),
            column.kind,
            column.width,
        )
        for column in _word_count_columns()
    ],
    ExportColumn("客户交稿时间", lambda pair: _read(pair[1], "customer_deadline_time"), "datetime", 20),
    ExportColumn("发客户时间", lambda pair: _read(pair[1], "sent_to_client_time"), "datetime", 20),
    ExportColumn("客户反馈", lambda pair: _read(pair[1], "client_feedback"), width=36),
    ExportColumn("已分配译员", lambda pair: _assigned_translators(pair[1]), width=28),
    ExportColumn("译员回稿时间", lambda pair: _translator_return_times(pair[1]), width=28),
    ExportColumn("译员任务完成情况", lambda pair: _translator_completion_remarks(pair[1]), width=36),
    ExportColumn("译员分配时间", lambda pair: _read(pair[1], "translator_assignment_time"), "datetime", 20),
    ExportColumn("译员交付进度", lambda pair: _read(pair[1], "translator_delivery_progress"), "percent", 16),
    ExportColumn("审校前 QC", lambda pair: _read(pair[1], "pre_review_qc_progress"), "percent", 14),
    ExportColumn("审核进度（旧字段）", lambda pair: _read(pair[1], "review_progress"), "percent", 18),
    ExportColumn("审校 1", lambda pair: _read(pair[1], "review1_progress"), "percent", 12),
    ExportColumn("审校 2", lambda pair: _read(pair[1], "review2_progress"), "percent", 12),
    ExportColumn("审校后 QC", lambda pair: _read(pair[1], "post_review_qc_progress"), "percent", 14),
    ExportColumn("排版进度", lambda pair: _read(pair[1], "layout_progress"), "percent", 14),
    ExportColumn("整合进度", lambda pair: _read(pair[1], "consolidation_progress"), "percent", 14),
    ExportColumn("网络文件路径", lambda pair: _read(pair[1], "network_file_path"), width=38),
    ExportColumn("备注", lambda pair: _read(pair[1], "remarks"), width=36),
    ExportColumn("创建时间", lambda pair: _read(pair[1], "created_at"), "datetime", 20),
    ExportColumn("更新时间", lambda pair: _read(pair[1], "updated_at"), "datetime", 20),
]


def _percent_value(value: Any) -> float | str | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value) / 100
    normalized = str(value).strip()
    try:
        return float(normalized.rstrip("%")) / 100
    except ValueError:
        return _safe_text(normalized)


def _cell_value(value: Any, kind: str) -> Any:
    if value is None:
        return None
    if kind == "percent":
        return _percent_value(value)
    if kind == "integer":
        try:
            return int(value)
        except (TypeError, ValueError):
            return _safe_text(value)
    if kind == "datetime" and isinstance(value, (date, datetime)):
        return value
    return _safe_text(value)


def _make_sheet(workbook: Workbook, title: str, columns: Sequence[ExportColumn]):
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.row_dimensions[1].height = 30
    for index, column in enumerate(columns, 1):
        sheet.column_dimensions[get_column_letter(index)].width = min(column.width, 40)
    headers = []
    for column in columns:
        cell = WriteOnlyCell(sheet, value=column.label)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        headers.append(cell)
    sheet.append(headers)
    return sheet


def _append_row(sheet, columns: Sequence[ExportColumn], item: Any) -> None:
    cells = []
    for column in columns:
        cell = WriteOnlyCell(sheet, value=_cell_value(column.getter(item), column.kind))
        cell.font = Font(name="微软雅黑", size=10, color="1F2937")
        cell.alignment = Alignment(
            horizontal="right" if column.kind in {"integer", "percent"} else "left",
            vertical="top",
            wrap_text=column.kind not in {"integer", "percent", "datetime"},
        )
        if column.kind == "datetime":
            cell.number_format = "yyyy-mm-dd hh:mm"
        elif column.kind == "integer":
            cell.number_format = "#,##0"
        elif column.kind == "percent":
            cell.number_format = "0%"
        elif column.kind == "identifier":
            cell.number_format = "@"
        cells.append(cell)
    sheet.append(cells)


def translation_projects_to_xlsx(
    batches: Iterable[Sequence[Any]],
    *,
    max_rows_per_sheet: int = EXPORT_MAX_ROWS_PER_SHEET,
) -> bytes:
    """把已挂载关联数据的项目批次写成双工作表 XLSX。"""
    workbook = Workbook(write_only=True)
    project_sheet = _make_sheet(workbook, "母订单", PROJECT_COLUMNS)
    sub_order_sheet = _make_sheet(workbook, "子订单", SUB_ORDER_COLUMNS)
    project_count = 0
    sub_order_count = 0

    try:
        for batch in batches:
            for project in batch:
                project_count += 1
                if project_count > max_rows_per_sheet:
                    raise TranslationExportLimitError(
                        f"母订单超过 {max_rows_per_sheet} 行，请缩小时间范围"
                    )
                _append_row(project_sheet, PROJECT_COLUMNS, project)
                sub_orders = sorted(
                    _read(project, "sub_orders", []) or [],
                    key=lambda item: str(_read(item, "sub_order_no") or ""),
                )
                for sub_order in sub_orders:
                    sub_order_count += 1
                    if sub_order_count > max_rows_per_sheet:
                        raise TranslationExportLimitError(
                            f"子订单超过 {max_rows_per_sheet} 行，请缩小时间范围"
                        )
                    _append_row(sub_order_sheet, SUB_ORDER_COLUMNS, (project, sub_order))

        if not project_count:
            raise TranslationExportEmptyError("所选范围内没有可导出的数据")
    except Exception:
        # write_only 工作表使用流式 XML writer；提前失败时保存到废弃缓冲区以完整关闭 writer。
        try:
            workbook.save(BytesIO())
        except Exception:
            pass
        raise

    project_sheet.auto_filter.ref = f"A1:{get_column_letter(len(PROJECT_COLUMNS))}{project_count + 1}"
    sub_order_sheet.auto_filter.ref = f"A1:{get_column_letter(len(SUB_ORDER_COLUMNS))}{sub_order_count + 1}"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _project_batches(
    db: Session,
    *,
    keyword: str | None,
    field_filters: dict,
    sort: str | None,
) -> Iterator[Sequence[Any]]:
    skip = 0
    while True:
        projects = get_translation_projects(
            db,
            skip=skip,
            limit=EXPORT_BATCH_SIZE,
            keyword=keyword,
            field_filters=field_filters,
            sort=sort,
        )
        if not projects:
            return
        yield projects
        if len(projects) < EXPORT_BATCH_SIZE:
            return
        skip += EXPORT_BATCH_SIZE


def create_translation_project_export(
    db: Session,
    *,
    keyword: str | None,
    field_filters: dict,
    sort: str | None,
) -> bytes:
    """复用列表查询分批读取全部命中项目并生成 XLSX。"""
    return translation_projects_to_xlsx(
        _project_batches(
            db,
            keyword=keyword,
            field_filters=field_filters,
            sort=sort,
        )
    )
