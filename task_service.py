"""个人任务、统一工作项和日报的业务逻辑。"""
from __future__ import annotations

import calendar
import datetime
from io import BytesIO
from uuid import UUID
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import AppUser, TranslationProject, TranslationSubOrder
from permission_registry import ALL_PERMISSION
from permission_service import get_user_permission_codes, user_has_permission
from leave_service import ensure_user_assignable
from task_models import (
    DailyReport,
    DailyReportItem,
    NonProjectTask,
    NonProjectTaskEvent,
    NonProjectTaskRecurrence,
    WorkEntry,
    TaskActivityEvent,
)
from task_schemas import (
    DailyReportItemInput,
    DailyReportSaveRequest,
    NonProjectTaskCreate,
    NonProjectTaskUpdate,
    RecurrenceCreate,
    WorkEntryCreate,
    WorkEntryUpdate,
)
from workflow_crud import get_my_tasks
from workflow_models import ProjectWorkbenchResponsibility, WorkflowInstance, WorkflowLog


BUSINESS_TZ = ZoneInfo("Asia/Hong_Kong")
OPEN_TASK_STATUSES = ("pending", "in_progress")


def business_now() -> datetime.datetime:
    """返回香港业务时区的无时区时间，与现有 TIMESTAMP 字段保持兼容。"""
    return datetime.datetime.now(BUSINESS_TZ).replace(tzinfo=None)


def normalize_business_datetime(value: datetime.datetime | None) -> datetime.datetime | None:
    if value is None or value.tzinfo is None:
        return value
    return value.astimezone(BUSINESS_TZ).replace(tzinfo=None)


def day_range(day: datetime.date) -> tuple[datetime.datetime, datetime.datetime]:
    start = datetime.datetime.combine(day, datetime.time.min)
    return start, start + datetime.timedelta(days=1)


def display_user(user: AppUser | None) -> str:
    if not user:
        return ""
    return user.full_name or user.username


def _get_active_user(db: Session, user_id: UUID) -> AppUser:
    user = db.query(AppUser).filter(AppUser.id == user_id, AppUser.is_active == True).first()
    if not user:
        raise LookupError("负责人不存在或已停用")
    return user


def _can_assign(db: Session, user_id: UUID) -> bool:
    return user_has_permission(db, user_id, "tasks:assign")


def _ensure_assignee_allowed(db: Session, operator_id: UUID, assignee_id: UUID) -> AppUser:
    user = _get_active_user(db, assignee_id)
    if assignee_id != operator_id and not _can_assign(db, operator_id):
        raise PermissionError("缺少任务分配权限，只能为自己创建任务")
    ensure_user_assignable(db, assignee_id)
    return user


def _task_event(
    db: Session,
    task: NonProjectTask,
    operator_id: UUID,
    event_type: str,
    *,
    from_status: str | None = None,
    to_status: str | None = None,
    detail: dict | None = None,
) -> None:
    db.add(
        NonProjectTaskEvent(
            task_id=task.id,
            operator_id=operator_id,
            event_type=event_type,
            from_status=from_status,
            to_status=to_status,
            detail=detail,
        )
    )


def serialize_non_project_task(task: NonProjectTask) -> dict:
    return {
        "id": task.id,
        "task_type": task.task_type,
        "task_name": task.task_name,
        "assigner_id": task.assigner_id,
        "assigner_name": display_user(task.assigner),
        "assignee_id": task.assignee_id,
        "assignee_name": display_user(task.assignee),
        "assigned_at": task.assigned_at,
        "planned_completion_at": task.planned_completion_at,
        "actual_completion_at": task.actual_completion_at,
        "status": task.status,
        "remark": task.remark,
        "recurrence_template_id": task.recurrence_template_id,
        "occurrence_date": task.occurrence_date,
        "created_at": task.created_at,
        "updated_at": task.updated_at,
    }


def create_non_project_task(
    db: Session, operator: AppUser, payload: NonProjectTaskCreate
) -> NonProjectTask:
    assignee_id = payload.assignee_id or operator.id
    _ensure_assignee_allowed(db, operator.id, assignee_id)
    now = business_now()
    task = NonProjectTask(
        task_type=payload.task_type.strip(),
        task_name=payload.task_name.strip(),
        assigner_id=operator.id,
        assignee_id=assignee_id,
        assigned_at=now,
        planned_completion_at=normalize_business_datetime(payload.planned_completion_at),
        status="pending",
        remark=payload.remark,
        created_at=now,
        updated_at=now,
    )
    db.add(task)
    db.flush()
    _task_event(db, task, operator.id, "created", to_status="pending")
    db.commit()
    return (
        db.query(NonProjectTask)
        .filter(NonProjectTask.id == task.id)
        .first()
    )


def list_non_project_tasks(
    db: Session,
    current_user: AppUser,
    *,
    include_created: bool = False,
    include_all: bool = False,
    status: str | None = None,
) -> list[NonProjectTask]:
    query = db.query(NonProjectTask)
    can_view_all = include_all and ALL_PERMISSION in get_user_permission_codes(db, current_user.id)
    if can_view_all:
        pass
    elif include_created:
        query = query.filter(
            (NonProjectTask.assignee_id == current_user.id)
            | (NonProjectTask.assigner_id == current_user.id)
        )
    else:
        query = query.filter(NonProjectTask.assignee_id == current_user.id)
    if status:
        query = query.filter(NonProjectTask.status == status)
    return query.order_by(
        NonProjectTask.planned_completion_at.asc().nullslast(),
        NonProjectTask.assigned_at.desc(),
    ).all()


def get_non_project_task(db: Session, task_id: UUID) -> NonProjectTask:
    task = db.query(NonProjectTask).filter(NonProjectTask.id == task_id).first()
    if not task:
        raise LookupError("非项目任务不存在")
    return task


def _ensure_task_visible(db: Session, task: NonProjectTask, user: AppUser) -> None:
    if task.assignee_id == user.id or task.assigner_id == user.id or _can_assign(db, user.id):
        return
    raise PermissionError("无权查看或操作该任务")


def update_non_project_task(
    db: Session, task_id: UUID, operator: AppUser, payload: NonProjectTaskUpdate
) -> NonProjectTask:
    task = get_non_project_task(db, task_id)
    _ensure_task_visible(db, task, operator)
    changes = payload.model_dump(exclude_unset=True)
    if "planned_completion_at" in changes:
        changes["planned_completion_at"] = normalize_business_datetime(
            changes["planned_completion_at"]
        )
    if "assignee_id" in changes:
        new_assignee = changes["assignee_id"]
        if new_assignee is None:
            raise ValueError("负责人不能为空")
        _ensure_assignee_allowed(db, operator.id, new_assignee)
        if new_assignee != task.assignee_id:
            changes["assigned_at"] = business_now()
    if task.status == "cancelled":
        raise ValueError("已取消任务不能编辑")
    for key, value in changes.items():
        if key in {"task_type", "task_name"} and isinstance(value, str):
            value = value.strip()
        setattr(task, key, value)
    task.updated_at = business_now()
    _task_event(db, task, operator.id, "updated", detail={"fields": list(changes)})
    db.commit()
    return get_non_project_task(db, task.id)


def change_task_status(
    db: Session, task_id: UUID, operator: AppUser, action: str, note: str | None
) -> NonProjectTask:
    task = get_non_project_task(db, task_id)
    _ensure_task_visible(db, task, operator)
    if task.assignee_id != operator.id and not _can_assign(db, operator.id):
        raise PermissionError("只有负责人或任务管理员可以变更状态")
    transitions = {
        "start": (("pending",), "in_progress"),
        "complete": (("pending", "in_progress"), "completed"),
        "reopen": (("completed",), "in_progress"),
        "cancel": (("pending", "in_progress", "completed"), "cancelled"),
    }
    if action not in transitions:
        raise ValueError("未知任务操作")
    allowed, target = transitions[action]
    if task.status not in allowed:
        raise ValueError(f"当前状态不能执行“{action}”操作")
    old_status = task.status
    task.status = target
    task.updated_at = business_now()
    task.actual_completion_at = business_now() if target == "completed" else None
    _task_event(
        db,
        task,
        operator.id,
        action,
        from_status=old_status,
        to_status=target,
        detail={"note": note} if note else None,
    )
    db.commit()
    return get_non_project_task(db, task.id)


def create_work_entry(
    db: Session, operator: AppUser, payload: WorkEntryCreate
) -> WorkEntry:
    if payload.non_project_task_id:
        task = get_non_project_task(db, payload.non_project_task_id)
        if task.assignee_id != operator.id:
            raise PermissionError("只能为自己负责的任务填写工作记录")
        if task.status == "pending":
            task.status = "in_progress"
            task.updated_at = business_now()
            _task_event(
                db,
                task,
                operator.id,
                "start",
                from_status="pending",
                to_status="in_progress",
            )
    elif payload.project_responsibility_id:
        responsibility = db.query(ProjectWorkbenchResponsibility).filter(
            ProjectWorkbenchResponsibility.id == payload.project_responsibility_id
        ).first()
        if not responsibility or responsibility.assignee_id != operator.id:
            raise PermissionError("只能为当前由自己负责的项目责任填写工作记录")
        from project_workbench_service import is_active_project
        if not responsibility.project or not is_active_project(
            responsibility.project_type, responsibility.project.project_status
        ):
            raise ValueError("该项目已不在工作台活跃范围")
    else:
        my_workflow_ids = {
            item["workflow_instance_id"] for item in get_my_tasks(db, operator.id)
        }
        if payload.workflow_instance_id not in my_workflow_ids:
            raise PermissionError("只能为当前由自己负责的项目任务填写工作记录")
    now = business_now()
    entry = WorkEntry(
        user_id=operator.id,
        work_date=payload.work_date,
        workflow_instance_id=payload.workflow_instance_id,
        project_responsibility_id=payload.project_responsibility_id,
        non_project_task_id=payload.non_project_task_id,
        progress_content=payload.progress_content.strip(),
        duration_minutes=payload.duration_minutes,
        result_content=payload.result_content,
        created_at=now,
        updated_at=now,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return entry


def update_work_entry(
    db: Session, entry_id: UUID, operator: AppUser, payload: WorkEntryUpdate
) -> WorkEntry:
    entry = (
        db.query(WorkEntry)
        .filter(WorkEntry.id == entry_id, WorkEntry.user_id == operator.id)
        .first()
    )
    if not entry:
        raise LookupError("工作记录不存在")
    for key, value in payload.model_dump(exclude_unset=True).items():
        if key == "progress_content" and isinstance(value, str):
            value = value.strip()
        setattr(entry, key, value)
    entry.updated_at = business_now()
    db.commit()
    db.refresh(entry)
    return entry


def list_work_entries(
    db: Session, user_id: UUID, work_date: datetime.date | None = None
) -> list[WorkEntry]:
    query = db.query(WorkEntry).filter(WorkEntry.user_id == user_id)
    if work_date:
        query = query.filter(WorkEntry.work_date == work_date)
    return query.order_by(WorkEntry.work_date.desc(), WorkEntry.created_at.desc()).all()


def create_recurrence(
    db: Session, operator: AppUser, payload: RecurrenceCreate
) -> NonProjectTaskRecurrence:
    assignee_id = payload.assignee_id or operator.id
    _ensure_assignee_allowed(db, operator.id, assignee_id)
    now = business_now()
    recurrence = NonProjectTaskRecurrence(
        task_type=payload.task_type.strip(),
        task_name=payload.task_name.strip(),
        assigner_id=operator.id,
        assignee_id=assignee_id,
        frequency=payload.frequency,
        weekdays=sorted(set(payload.weekdays or [])) or None,
        month_day=payload.month_day,
        default_due_time=payload.default_due_time,
        start_date=payload.start_date,
        end_date=payload.end_date,
        remark=payload.remark,
        is_active=True,
        created_at=now,
        updated_at=now,
    )
    db.add(recurrence)
    db.commit()
    db.refresh(recurrence)
    today = business_now().date()
    generate_recurrence_instances(db, min(today, payload.end_date or today), recurrence.id)
    return recurrence


def _matches_recurrence(rule: NonProjectTaskRecurrence, day: datetime.date) -> bool:
    if rule.frequency == "daily":
        return True
    if rule.frequency == "workday":
        return day.weekday() < 5
    if rule.frequency == "weekly":
        weekdays = rule.weekdays or [rule.start_date.weekday()]
        return day.weekday() in weekdays
    target = rule.month_day or rule.start_date.day
    target = min(target, calendar.monthrange(day.year, day.month)[1])
    return day.day == target


def generate_recurrence_instances(
    db: Session,
    through_date: datetime.date,
    recurrence_id: UUID | None = None,
    assignee_id: UUID | None = None,
) -> int:
    if through_date > business_now().date() + datetime.timedelta(days=366):
        raise ValueError("周期任务最多生成到未来一年")
    query = db.query(NonProjectTaskRecurrence).filter(
        NonProjectTaskRecurrence.is_active == True,
        NonProjectTaskRecurrence.start_date <= through_date,
    )
    if recurrence_id:
        query = query.filter(NonProjectTaskRecurrence.id == recurrence_id)
    if assignee_id:
        query = query.filter(NonProjectTaskRecurrence.assignee_id == assignee_id)
    created = 0
    for rule in query.all():
        end_date = min(through_date, rule.end_date or through_date)
        existing_dates = {
            row.occurrence_date
            for row in db.query(NonProjectTask.occurrence_date)
            .filter(
                NonProjectTask.recurrence_template_id == rule.id,
                NonProjectTask.occurrence_date >= rule.start_date,
                NonProjectTask.occurrence_date <= end_date,
            )
            .all()
        }
        day = rule.start_date
        while day <= end_date:
            if day not in existing_dates and _matches_recurrence(rule, day):
                due_at = (
                    datetime.datetime.combine(day, rule.default_due_time)
                    if rule.default_due_time
                    else None
                )
                task = NonProjectTask(
                    task_type=rule.task_type,
                    task_name=rule.task_name,
                    assigner_id=rule.assigner_id,
                    assignee_id=rule.assignee_id,
                    assigned_at=business_now(),
                    planned_completion_at=due_at,
                    status="pending",
                    remark=rule.remark,
                    recurrence_template_id=rule.id,
                    occurrence_date=day,
                    created_at=business_now(),
                    updated_at=business_now(),
                )
                db.add(task)
                created += 1
            day += datetime.timedelta(days=1)
    db.commit()
    return created


def get_my_work_items(db: Session, current_user: AppUser) -> list[dict]:
    from workflow_delegation_service import close_completed_delegations, notify_overdue_delegations
    close_completed_delegations(db)
    notify_overdue_delegations(db)
    is_super_admin = ALL_PERMISSION in get_user_permission_codes(db, current_user.id)
    generate_recurrence_instances(
        db,
        business_now().date(),
        assignee_id=None if is_super_admin else current_user.id,
    )
    items: list[dict] = []
    for task in get_my_tasks(db, current_user.id, include_all=is_super_admin):
        project_source_id = task.get("workflow_instance_id") or task.get("project_responsibility_id")
        assignment_type = task.get("assignment_type")
        available_actions = ["enter_project"]
        if assignment_type == "direct":
            available_actions.extend(["work_entry", "handover"])
        if task.get("delegation_id") and assignment_type in {"direct", "delegated_out"}:
            available_actions.append("return_delegation")
        items.append(
            {
                "source_type": "project",
                "source_id": project_source_id,
                "task_type": task.get("task_type") or "项目任务",
                "task_name": task["sub_project_name"] or task["project_name"],
                "assigner_name": None,
                "assigned_at": None,
                "planned_completion_at": task["customer_deadline_time"],
                "actual_completion_at": None,
                "status": task["project_status"] or "pending",
                "remark": None,
                "available_actions": available_actions,
                **task,
            }
        )
    can_assign = _can_assign(db, current_user.id)
    for task in list_non_project_tasks(
        db,
        current_user,
        include_created=can_assign,
        include_all=is_super_admin,
    ):
        serialized = serialize_non_project_task(task)
        can_change_status = task.assignee_id == current_user.id or can_assign
        available_actions = []
        if task.status == "pending" and can_change_status:
            available_actions.append("start")
        if (
            task.status in OPEN_TASK_STATUSES
            and task.assignee_id == current_user.id
        ):
            available_actions.append("work_entry")
        if task.status in OPEN_TASK_STATUSES and can_change_status:
            available_actions.append("complete")
        if task.status != "cancelled":
            available_actions.append("edit")
        if task.status in (*OPEN_TASK_STATUSES, "completed") and can_change_status:
            available_actions.append("cancel")
        if task.status == "completed" and can_change_status:
            available_actions.append("reopen")
        items.append(
            {
                "source_type": "non_project",
                "source_id": task.id,
                "task_type": task.task_type,
                "task_name": task.task_name,
                "assigner_name": serialized["assigner_name"],
                "assigner_id": task.assigner_id,
                "assignee_id": task.assignee_id,
                "assignee_name": serialized["assignee_name"],
                "assigned_at": task.assigned_at,
                "planned_completion_at": task.planned_completion_at,
                "actual_completion_at": task.actual_completion_at,
                "status": task.status,
                "remark": task.remark,
                "assignment_type": (
                    "overview"
                    if is_super_admin and task.assignee_id != current_user.id
                    else "direct"
                ),
                "available_actions": available_actions,
            }
        )
    items.sort(
        key=lambda item: (
            item["planned_completion_at"] is None,
            item["planned_completion_at"] or datetime.datetime.max,
            item["source_type"],
        )
    )
    return items


def _project_identity(db: Session, workflow_id: UUID) -> tuple[str, str, dict]:
    wf = db.query(WorkflowInstance).filter(WorkflowInstance.id == workflow_id).first()
    if not wf:
        return "项目任务", "已删除的项目任务", {}
    if wf.sub_order_id:
        sub = db.query(TranslationSubOrder).filter(TranslationSubOrder.id == wf.sub_order_id).first()
        project = (
            db.query(TranslationProject)
            .filter(TranslationProject.id == sub.parent_project_id)
            .first()
            if sub
            else None
        )
        return (
            (project.task_type or "项目任务") if project else "项目任务",
            (sub.sub_project_name or sub.sub_order_no) if sub else "子订单任务",
            {
                "order_no": sub.sub_order_no if sub else None,
                "project_name": project.project_name if project else None,
                "client_short_name": project.client.client_short_name if project and project.client else None,
                "workflow_instance_id": str(workflow_id),
            },
        )
    project = (
        db.query(TranslationProject)
        .filter(TranslationProject.id == wf.translation_project_id)
        .first()
    )
    return (
        (project.task_type or "项目任务") if project else "项目任务",
        (project.project_name or project.order_no) if project else "项目任务",
        {
            "order_no": project.order_no if project else None,
            "project_name": project.project_name if project else None,
            "client_short_name": project.client.client_short_name if project and project.client else None,
            "workflow_instance_id": str(workflow_id),
        },
    )


def _responsibility_identity(db: Session, responsibility_id: UUID) -> tuple[str, str, dict]:
    from project_workbench_service import serialize_responsibility
    row = db.query(ProjectWorkbenchResponsibility).filter(
        ProjectWorkbenchResponsibility.id == responsibility_id
    ).first()
    if not row or not row.project:
        return "项目任务", "已删除的项目责任", {}
    item = serialize_responsibility(db, row)
    return (
        item["task_type"],
        item["project_name"],
        {
            "order_no": item["order_no"],
            "project_name": item["project_name"],
            "client_short_name": item["client_short_name"],
            "project_type": item["project_type"],
            "project_type_label": item["project_type_label"],
            "responsibility_role_code": item["current_stage_role_code"],
            "responsibility_role_name": item["current_stage_role_name"],
            "project_responsibility_id": str(responsibility_id),
        },
    )
def derive_daily_report_items(
    db: Session, user_id: UUID, report_date: datetime.date
) -> list[dict]:
    items: list[dict] = []
    entries = list_work_entries(db, user_id, report_date)
    non_project_with_entry: set[UUID] = set()
    for entry in entries:
        if entry.non_project_task_id:
            task = get_non_project_task(db, entry.non_project_task_id)
            source_type, task_type, task_name = "non_project", task.task_type, task.task_name
            metadata = {"work_entry_id": str(entry.id)}
            non_project_with_entry.add(task.id)
            source_id = task.id
        elif entry.project_responsibility_id:
            task_type, task_name, metadata = _responsibility_identity(
                db, entry.project_responsibility_id
            )
            source_type = "project"
            source_id = entry.project_responsibility_id
            metadata["work_entry_id"] = str(entry.id)
        else:
            task_type, task_name, metadata = _project_identity(
                db, entry.workflow_instance_id
            )
            source_type = "project"
            source_id = entry.workflow_instance_id
            metadata["work_entry_id"] = str(entry.id)
        items.append(
            {
                "source_type": source_type,
                "source_id": source_id,
                "task_type": task_type,
                "task_name": task_name,
                "progress_content": entry.progress_content,
                "result_content": entry.result_content,
                "duration_minutes": entry.duration_minutes,
                "display_metadata": metadata,
            }
        )

    start, end = day_range(report_date)
    completed = (
        db.query(NonProjectTask)
        .filter(
            NonProjectTask.assignee_id == user_id,
            NonProjectTask.actual_completion_at >= start,
            NonProjectTask.actual_completion_at < end,
        )
        .all()
    )
    for task in completed:
        if task.id in non_project_with_entry:
            continue
        items.append(
            {
                "source_type": "non_project",
                "source_id": task.id,
                "task_type": task.task_type,
                "task_name": task.task_name,
                "progress_content": "完成任务",
                "result_content": task.remark,
                "duration_minutes": 0,
                "display_metadata": {
                    "actual_completion_at": task.actual_completion_at.isoformat()
                },
            }
        )

    logs = (
        db.query(WorkflowLog)
        .filter(
            WorkflowLog.operator_id == user_id,
            WorkflowLog.created_at >= start,
            WorkflowLog.created_at < end,
        )
        .order_by(WorkflowLog.created_at)
        .all()
    )
    for log in logs:
        if log.direction == "handover":
            # 交接由双方独立的系统活动事件进入日报，避免原负责人重复出现两行。
            continue
        task_type, task_name, metadata = _project_identity(
            db, log.workflow_instance_id
        )
        metadata.update(
            {
                "workflow_log_id": str(log.id),
                "from_stage": log.from_stage,
                "to_stage": log.to_stage,
            }
        )
        items.append(
            {
                "source_type": "project",
                "source_id": log.workflow_instance_id,
                "task_type": task_type,
                "task_name": task_name,
                "progress_content": log.description
                or f"工作流从 {log.from_stage or '-'} 流转至 {log.to_stage or '-'}",
                "result_content": log.note,
                "duration_minutes": 0,
                "display_metadata": metadata,
            }
        )

    events = db.query(TaskActivityEvent).filter(
        TaskActivityEvent.user_id == user_id,
        TaskActivityEvent.occurred_at >= start,
        TaskActivityEvent.occurred_at < end,
    ).order_by(TaskActivityEvent.occurred_at, TaskActivityEvent.id).all()
    if events:
        from task_activity_service import activity_to_report_item
        items.extend(activity_to_report_item(event) for event in events)
    return items


def _serialize_report(report: DailyReport) -> dict:
    return {
        "id": report.id,
        "user_id": report.user_id,
        "user_name": display_user(report.user),
        "report_date": report.report_date,
        "status": report.status,
        "supplemental_note": report.supplemental_note,
        "generated_at": report.generated_at,
        "finalized_at": report.finalized_at,
        "items": [
            {
                "id": item.id,
                "source_type": item.source_type,
                "source_id": item.source_id,
                "task_type": item.task_type,
                "task_name": item.task_name,
                "progress_content": item.progress_content,
                "result_content": item.result_content,
                "duration_minutes": item.duration_minutes,
                "display_metadata": item.display_metadata,
                "sort_order": item.sort_order,
            }
            for item in report.items
        ],
    }


def preview_daily_report(
    db: Session, user: AppUser, report_date: datetime.date, *, refresh: bool = False
) -> dict:
    report = (
        db.query(DailyReport)
        .filter(DailyReport.user_id == user.id, DailyReport.report_date == report_date)
        .first()
    )
    if report and (report.status == "finalized" or not refresh):
        return _serialize_report(report)
    if report:
        derived_items = derive_daily_report_items(db, user.id, report_date)
        return {
            "id": report.id,
            "user_id": report.user_id,
            "user_name": display_user(user),
            "report_date": report.report_date,
            "status": "draft",
            "supplemental_note": report.supplemental_note,
            "generated_at": business_now(),
            "finalized_at": None,
            "items": [
                {**item, "id": None, "sort_order": index}
                for index, item in enumerate(derived_items)
            ],
        }
    return {
        "id": None,
        "user_id": user.id,
        "user_name": display_user(user),
        "report_date": report_date,
        "status": "draft",
        "supplemental_note": None,
        "generated_at": business_now(),
        "finalized_at": None,
        "items": [
            {**item, "id": None, "sort_order": index}
            for index, item in enumerate(
                derive_daily_report_items(db, user.id, report_date)
            )
        ],
    }


def save_daily_report(
    db: Session,
    user: AppUser,
    report_date: datetime.date,
    payload: DailyReportSaveRequest,
    *,
    finalize: bool = False,
) -> DailyReport:
    report = (
        db.query(DailyReport)
        .filter(DailyReport.user_id == user.id, DailyReport.report_date == report_date)
        .first()
    )
    if report and report.status == "finalized":
        raise ValueError("日报已确认，不能继续修改或重复确认")
    now = business_now()
    if not report:
        report = DailyReport(
            user_id=user.id,
            report_date=report_date,
            status="draft",
            generated_at=now,
            created_at=now,
            updated_at=now,
        )
        db.add(report)
        db.flush()
    report.supplemental_note = payload.supplemental_note
    if payload.items is not None:
        raw_items = merge_daily_report_items(
            [item.model_dump() for item in payload.items],
            derive_daily_report_items(db, user.id, report_date),
        )
    else:
        raw_items = derive_daily_report_items(db, user.id, report_date)
    report.items.clear()
    db.flush()
    for index, item in enumerate(raw_items):
        report.items.append(
            DailyReportItem(report_id=report.id, sort_order=index, **item)
        )
    report.status = "finalized" if finalize else "draft"
    report.finalized_at = now if finalize else None
    report.updated_at = now
    db.commit()
    return (
        db.query(DailyReport)
        .filter(DailyReport.id == report.id)
        .first()
    )


def merge_daily_report_items(client_items: list[dict], derived_items: list[dict]) -> list[dict]:
    """保留用户内容，但用数据库派生事件替换全部客户端系统事件。"""
    result = [item for item in client_items if item.get("source_type") != "system_event"]
    result.extend(
        {**item, "duration_minutes": 0}
        for item in derived_items
        if item.get("source_type") == "system_event"
    )
    return result


def withdraw_daily_report(
    db: Session, user: AppUser, report_date: datetime.date
) -> DailyReport:
    from daily_report_mail_models import DailyReportMailDelivery

    report = db.query(DailyReport).filter(
        DailyReport.user_id == user.id,
        DailyReport.report_date == report_date,
    ).with_for_update().first()
    if not report:
        raise LookupError("日报不存在")
    if report.status != "finalized":
        raise ValueError("只有已确认日报可以撤回")
    sent = db.query(DailyReportMailDelivery.id).filter(
        DailyReportMailDelivery.report_id == report.id,
        DailyReportMailDelivery.status == "sent",
    ).first()
    if sent:
        raise ValueError("日报邮件已经发送，不能撤回或修改")
    report.status = "draft"
    report.finalized_at = None
    report.updated_at = business_now()
    db.commit()
    return report


def report_to_xlsx(report: DailyReport) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "个人日报"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

    sheet.merge_cells("A1:G1")
    sheet["A1"] = f"{display_user(report.user)}个人工作日报"
    sheet["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    sheet["A2"] = "日期"
    sheet["B2"] = report.report_date
    sheet["B2"].number_format = "yyyy-mm-dd"
    sheet["D2"] = "状态"
    sheet["E2"] = "已确认" if report.status == "finalized" else "草稿"
    sheet["A3"] = "补充说明"
    sheet.merge_cells("B3:G3")
    sheet["B3"] = report.supplemental_note or ""

    headers = ["任务类型", "任务名称", "工作进展", "工作成果", "耗时（分钟）", "项目/订单", "来源"]
    for col, value in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=col, value=value)
        cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    source_labels = {"project": "项目任务", "non_project": "非项目任务", "manual": "手工补充"}
    for row_index, item in enumerate(report.items, 5):
        metadata = item.display_metadata or {}
        order_info = metadata.get("order_no") or metadata.get("project_name") or ""
        values = [
            item.task_type,
            item.task_name,
            item.progress_content,
            item.result_content or "",
            item.duration_minutes,
            order_info,
            source_labels.get(item.source_type, item.source_type),
        ]
        for col, value in enumerate(values, 1):
            cell = sheet.cell(row=row_index, column=col, value=value)
            cell.font = Font(name="微软雅黑", size=10)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=col in {2, 3, 4},
                horizontal="right" if col == 5 else "left",
            )
        sheet.cell(row=row_index, column=5).number_format = "0"

    total_minutes = sum(max(0, int(item.duration_minutes or 0)) for item in report.items)
    total_hours = f"{total_minutes / 60:.2f}".rstrip("0").rstrip(".")
    summary_row = 5 + len(report.items)
    sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    sheet.cell(row=summary_row, column=1, value="当日工作耗时合计")
    sheet.cell(row=summary_row, column=5, value=f"{total_hours} 小时（{total_minutes} 分钟）")
    sheet.merge_cells(start_row=summary_row, start_column=6, end_row=summary_row, end_column=7)
    for col in range(1, 8):
        cell = sheet.cell(row=summary_row, column=col)
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="166534")
        cell.fill = PatternFill("solid", fgColor="ECFDF5")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[summary_row].height = 28

    widths = [14, 28, 38, 32, 14, 22, 14]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = f"A4:G{max(4, 4 + len(report.items))}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def serialize_report(report: DailyReport) -> dict:
    return _serialize_report(report)
