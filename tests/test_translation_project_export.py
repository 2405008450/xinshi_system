from datetime import date, datetime
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

import routers.translation_projects as translation_router
import translation_project_export_service as export_service


def sample_project(*, project_name="示例项目", with_sub_order=True):
    assignment = {
        "translator_name": "张三",
        "translation_scope": "正文",
        "translator_return_time": datetime(2026, 9, 5, 18, 30),
        "completion_remarks": "已完成",
    }
    sub_order = SimpleNamespace(
        sub_order_no="TP-260901-001.001",
        sub_project_name="子项目一",
        status="translator_assigned",
        word_count_matrix={
            "company": {"words": 500},
            "customer": {},
            "translator_estimate": {},
        },
        customer_deadline_time=datetime(2026, 9, 6, 17, 0),
        assigned_translators=[assignment],
        translator_delivery_progress="25%",
        remarks="=HYPERLINK(\"https://example.com\")",
    )
    return SimpleNamespace(
        order_no="TP-260901-001",
        project_name=project_name,
        quotation_required=True,
        project_status="confirmed",
        role_assignments=[
            {"role_code": "project_specialist", "assignee_name": "李四"},
            {"role_code": "project_assistant", "assignee_name": None},
        ],
        word_count_matrix={
            "company": {"words": 1000, "characters_no_spaces": 1200},
            "customer": {"pages": 8},
            "translator_estimate": {"foreign_words": 900},
        },
        customer_reception_time=datetime(2026, 9, 1, 9, 0),
        customer_deadline_time=datetime(2026, 9, 8, 18, 0),
        assigned_translators=[assignment],
        translator_delivery_progress="50%",
        sub_orders=[sub_order] if with_sub_order else [],
    )


def row_by_headers(sheet, row_number=2):
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[row_number]]
    return headers, dict(zip(headers, values))


def test_export_workbook_contains_complete_typed_project_and_sub_order_data():
    content = export_service.translation_projects_to_xlsx([[sample_project()]])
    workbook = load_workbook(BytesIO(content), data_only=False)

    assert workbook.sheetnames == ["母订单", "子订单"]
    project_headers, project_row = row_by_headers(workbook["母订单"])
    sub_headers, sub_row = row_by_headers(workbook["子订单"])

    assert sum(header.startswith(("我司-", "客户-", "译员预估-")) for header in project_headers) == 18
    assert project_row["订单号"] == "TP-260901-001"
    assert project_row["项目专员"] == "李四"
    assert project_row["项目助理"] == "角色池"
    assert project_row["状态"] == "已确认"
    assert project_row["我司-字数"] == 1000
    assert project_row["客户-页数"] == 8
    assert isinstance(project_row["客户接单时间"], datetime)
    assert project_row["译员交付进度"] == 0.5
    assert project_row["已分配译员"] == "张三（正文）"
    assert "张三：2026-09-05 18:30" in project_row["译员回稿时间"]

    assert sub_headers[:3] == ["母订单号", "母项目名称", "子订单号"]
    assert sub_row["母订单号"] == "TP-260901-001"
    assert sub_row["子订单号"] == "TP-260901-001.001"
    assert sub_row["我司-字数"] == 500
    assert sub_row["译员交付进度"] == 0.25

    remarks_cell = workbook["子订单"].cell(row=2, column=sub_headers.index("备注") + 1)
    assert remarks_cell.data_type == "s"
    assert remarks_cell.value.startswith("'=")
    assert workbook["母订单"].freeze_panes == "A2"
    assert workbook["子订单"].auto_filter.ref.endswith("2")


def test_export_rejects_empty_data_and_sheet_row_limit():
    with pytest.raises(export_service.TranslationExportEmptyError, match="没有可导出"):
        export_service.translation_projects_to_xlsx([])

    with pytest.raises(export_service.TranslationExportLimitError, match="母订单"):
        export_service.translation_projects_to_xlsx(
            [[sample_project(with_sub_order=False), sample_project(with_sub_order=False)]],
            max_rows_per_sheet=1,
        )


def test_export_filters_override_same_time_field_and_preserve_other_filters():
    raw = (
        '{"project_status":{"op":"in","value":["confirmed"]},'
        '"customer_reception_time":{"op":"between","from":"2026-08-01","to":"2026-08-31"},'
        '"customer_deadline_time":{"op":"between","from":"2026-09-15","to":"2026-09-30"}}'
    )
    filters = translation_router._export_field_filters(
        raw,
        time_field="customer_reception_time",
        date_start=date(2026, 9, 1),
        date_end=date(2026, 9, 10),
    )

    assert filters["project_status"]["value"] == ["confirmed"]
    assert filters["customer_reception_time"] == {
        "op": "between", "from": "2026-09-01", "to": "2026-09-10",
    }
    assert filters["customer_deadline_time"]["from"] == "2026-09-15"


def test_export_filters_reject_reverse_date_range():
    with pytest.raises(HTTPException) as exc_info:
        translation_router._export_field_filters(
            None,
            time_field="created_at",
            date_start=date(2026, 9, 2),
            date_end=date(2026, 9, 1),
        )
    assert exc_info.value.status_code == 422
    assert exc_info.value.detail == "开始日期不能晚于结束日期"


def test_export_query_is_paged_and_reuses_all_filters(monkeypatch):
    calls = []
    projects = [sample_project(with_sub_order=False), sample_project(with_sub_order=False)]

    def fake_get_projects(_db, **kwargs):
        calls.append(kwargs)
        return projects[kwargs["skip"]:kwargs["skip"] + kwargs["limit"]]

    monkeypatch.setattr(export_service, "EXPORT_BATCH_SIZE", 1)
    monkeypatch.setattr(export_service, "get_translation_projects", fake_get_projects)
    content = export_service.create_translation_project_export(
        object(),
        keyword="客户A",
        field_filters={"created_at": {"op": "between", "from": "2026-09-01", "to": "2026-09-02"}},
        sort="order_no_desc",
    )

    assert content.startswith(b"PK")
    assert [call["skip"] for call in calls] == [0, 1, 2]
    assert all(call["keyword"] == "客户A" for call in calls)
    assert all(call["sort"] == "order_no_desc" for call in calls)


def test_export_route_returns_xlsx_headers(monkeypatch):
    monkeypatch.setattr(
        translation_router,
        "create_translation_project_export",
        lambda *_args, **_kwargs: b"xlsx-content",
    )
    response = translation_router.export_projects(
        time_field="created_at",
        date_start=date(2026, 9, 1),
        date_end=date(2026, 9, 30),
        keyword=None,
        sort=None,
        field_filters=None,
        db=object(),
    )

    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    disposition = response.headers["content-disposition"]
    assert "translation-projects-2026-09-01-2026-09-30.xlsx" in disposition
    assert "filename*=UTF-8''" in disposition


@pytest.mark.parametrize(
    ("exception", "status_code"),
    [
        (export_service.TranslationExportEmptyError("无数据"), 404),
        (export_service.TranslationExportLimitError("超限"), 422),
    ],
)
def test_export_route_maps_business_errors(monkeypatch, exception, status_code):
    def raise_error(*_args, **_kwargs):
        raise exception

    monkeypatch.setattr(translation_router, "create_translation_project_export", raise_error)
    with pytest.raises(HTTPException) as exc_info:
        translation_router.export_projects(
            time_field="created_at",
            date_start=date(2026, 9, 1),
            date_end=date(2026, 9, 30),
            keyword=None,
            sort=None,
            field_filters=None,
            db=object(),
        )
    assert exc_info.value.status_code == status_code
