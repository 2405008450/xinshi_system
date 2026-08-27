from io import BytesIO

import pytest
from openpyxl import Workbook

import annotation_account_import_service as import_service
import annotation_ops_service as account_service
from annotation_custom_field_service import PROJECT_SCOPED_TABLES
from annotation_ops_models import AnnotationAccountAssignment, AnnotationCustomFieldDefinition


def _workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "澳鹏int英语采集账号"
    sheet.append(["说明", "还未开始任务的", None, "未捕颜色：首次未提交"])
    sheet.append(["数据编号", "user email", "initial password", "所分配人员姓名", "性别", "错误点/问题", "错误点/问题", "价格"])
    sheet.append([1, "worker@example.com", "secret", "测试人员", "女", "一审反馈", "二审反馈", 60])
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def test_detects_second_header_row_and_disambiguates_repeated_columns():
    workbook = import_service._load_workbook(_workbook_bytes())
    sheet = workbook[workbook.sheetnames[0]]

    header_row = import_service._detect_header_row(sheet)
    headers = import_service._headers(sheet, header_row)
    mapping = import_service._normalize_mapping(headers, None)

    assert header_row == 2
    assert [item["uniqueLabel"] for item in headers if item["label"] == "错误点/问题"] == [
        "错误点/问题", "错误点/问题（第2列）",
    ]
    error_rules = [item for item in mapping if item.get("fieldKey", "").startswith("error_feedback")]
    assert [item["fieldKey"] for item in error_rules] == ["error_feedback", "error_feedback_2"]
    assert next(item for item in mapping if item["target"] == "login_account")["index"] == 1
    assert next(item for item in mapping if item.get("fieldKey") == "price")["dataType"] == "number"


def test_parses_rows_without_importing_instruction_row():
    workbook = import_service._load_workbook(_workbook_bytes())
    sheet = workbook[workbook.sheetnames[0]]
    headers = import_service._headers(sheet, 2)
    mapping = import_service._normalize_mapping(headers, None)

    rows = import_service._parse_rows(sheet, 2, headers, mapping)

    assert len(rows) == 1
    assert rows[0]["fixed"]["login_account"] == "worker@example.com"
    assert rows[0]["fixed"]["person_name"] == "测试人员"
    assert rows[0]["custom"]["price"] == 60


def test_select_project_field_requires_options_during_preview():
    headers = [{"index": 0, "label": "质检状态", "uniqueLabel": "质检状态", "suggestion": {"target": "ignore"}}]
    with pytest.raises(ValueError, match="选择项不能为空"):
        import_service._normalize_mapping(headers, [{
            "index": 0, "target": "new_custom", "fieldKey": "quality_status",
            "fieldLabel": "质检状态", "dataType": "single_select", "options": [],
        }])


def test_rejects_files_larger_than_ten_megabytes():
    with pytest.raises(ValueError, match="10MB"):
        import_service._load_workbook(b"x" * (import_service.MAX_FILE_SIZE + 1))


def test_project_account_fields_are_scoped_and_assignment_keeps_values():
    assert "account_assignment" in PROJECT_SCOPED_TABLES
    assert AnnotationAccountAssignment.__table__.c.custom_values.nullable is False
    table_constraint = next(
        item for item in AnnotationCustomFieldDefinition.__table__.constraints
        if item.name == "ck_annotation_custom_field_table"
    )
    assert "account_assignment" in str(table_constraint.sqltext)


def test_account_login_masking_preserves_domain_but_not_secret():
    masked = account_service._mask_login_account("worker123@example.com")
    assert masked.endswith("@example.com")
    assert "worker123" not in masked
    assert account_service._mask_login_account("abcd") == "****"
