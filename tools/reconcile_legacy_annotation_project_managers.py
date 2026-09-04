"""把旧台账“项目部”字段对齐到标注项目的项目经理责任关系。

默认只生成预览报告。仅当一条记录中能找到唯一、启用且具备“项目经理”角色的
系统用户时才补齐；已有项目经理不覆盖，无法唯一确认的原值保留给人工处理。
"""

from __future__ import annotations

import argparse
import json
import re
import socket
import sys
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import annotation_ops_service  # noqa: F401  注册 SQLAlchemy 关联模型
import annotation_service  # noqa: F401  注册 SQLAlchemy 关联模型
from annotation_models import AnnotationProject
from database import SessionLocal
from leave_service import ensure_user_assignable
from models import AppUser
from workflow_models import ProjectWorkbenchResponsibility


ALIASES = {
    "景瀚": "肖景瀚",
    "静玲": "陈静玲",
    "欧阳婧琳": "欧阳靖琳",
}


def normalize(value) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().casefold()
    return "".join(re.findall(r"[0-9a-z\u4e00-\u9fff]+", text))


def source_project_department(path: Path) -> dict[int, str]:
    worksheet = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    return {
        row_number: str(worksheet.cell(row_number, 5).value or "").strip()
        for row_number in range(2, worksheet.max_row + 1)
    }


def split_people(raw_value: str) -> list[str]:
    return [
        value.strip()
        for value in re.split(r"[\n\r、,，/&＋+]", raw_value or "")
        if value.strip() and value.strip() not in {"无", "暂无"}
    ]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", help="旧标注项目 XLSX 文件路径")
    parser.add_argument("--report", required=True, help="JSON 报告输出路径")
    parser.add_argument("--apply", action="store_true", help="实际更新；默认仅预览")
    parser.add_argument("--expected-host", default="WIN-LOLJ8UHT2G5")
    args = parser.parse_args()

    if args.apply and socket.gethostname().casefold() != args.expected_host.casefold():
        raise RuntimeError(f"主机校验失败：当前为 {socket.gethostname()}，预期为 {args.expected_host}")

    raw_by_row = source_project_department(Path(args.source).resolve())
    db = SessionLocal()
    report = {
        "mode": "apply" if args.apply else "dry-run",
        "host": socket.gethostname(),
        "updated": [],
        "would_update": [],
        "already_assigned": [],
        "unmatched": [],
        "ambiguous": [],
        "failed": [],
    }
    try:
        # 本次旧台账整理按用户主数据中的部门归属匹配，部门是用户已人工确认的岗位口径。
        candidates = db.query(AppUser).filter(
            AppUser.is_active.is_(True),
            AppUser.department == "项目经理",
        ).all()
        candidate_by_key = {}
        for user in candidates:
            for label in (user.full_name, user.username):
                key = normalize(label)
                if key:
                    candidate_by_key.setdefault(key, user)

        projects = db.query(AnnotationProject).filter(
            AnnotationProject.idempotency_key.like("legacy-annotation-%")
        ).all()
        responsibilities = db.query(ProjectWorkbenchResponsibility).filter(
            ProjectWorkbenchResponsibility.annotation_project_id.in_([item.id for item in projects]),
            ProjectWorkbenchResponsibility.role_code == "project_manager",
        ).all()
        responsibility_by_project = {
            item.annotation_project_id: item for item in responsibilities
        }

        for project in projects:
            row_number = int(project.idempotency_key.rsplit("-", 1)[-1])
            raw_value = raw_by_row.get(row_number, "")
            responsibility = responsibility_by_project.get(project.id)
            base = {
                "row": row_number,
                "order_no": project.order_no,
                "project_name": project.project_name,
                "source_project_department": raw_value,
            }
            if responsibility is None:
                report["failed"].append({**base, "error": "缺少项目经理责任关系记录"})
                continue
            if responsibility.assignee_id:
                report["already_assigned"].append({
                    **base,
                    "manager": responsibility.assignee.full_name or responsibility.assignee.username,
                })
                continue

            matched = []
            for raw_name in split_people(raw_value):
                alias = ALIASES.get(raw_name, raw_name)
                user = candidate_by_key.get(normalize(alias))
                if user and user.id not in {item.id for item in matched}:
                    matched.append(user)
            if not matched:
                report["unmatched"].append(base)
                continue
            user = matched[0]
            target = {**base, "manager_id": str(user.id), "manager": user.full_name or user.username}
            if len(matched) > 1:
                target["other_matched_managers"] = [
                    item.full_name or item.username for item in matched[1:]
                ]
            try:
                ensure_user_assignable(db, user.id)
                if not args.apply:
                    report["would_update"].append(target)
                    continue
                responsibility.assignee_id = user.id
                responsibility.updated_at = datetime.now()
                db.commit()
                report["updated"].append(target)
            except Exception as exc:
                db.rollback()
                report["failed"].append({**target, "error": str(exc)})

        unresolved_counter = Counter(
            item["source_project_department"]
            for category in ("unmatched", "ambiguous")
            for item in report[category]
        )
        report["unresolved_values"] = [
            {"value": value, "count": count}
            for value, count in unresolved_counter.most_common()
        ]
        report["summary"] = {
            key: len(report[key])
            for key in ("updated", "would_update", "already_assigned", "unmatched", "ambiguous", "failed")
        }
    finally:
        db.close()

    report_path = Path(args.report).resolve()
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
