"""将旧标注项目台账导入当前标注项目模块。

默认仅生成预览报告；显式传入 ``--apply`` 后才写入数据库。脚本通过源文件哈希、
行号、项目名称、项目路径以及同客户同日期的近似名称进行幂等去重。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import socket
import sys
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from sqlalchemy import func, or_

# 允许从 tools 目录直接执行，同时加载项目根目录模块。
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import annotation_ops_service  # noqa: F401  注册 SQLAlchemy 关联模型
import annotation_service
from annotation_models import AnnotationProject
from annotation_ops_models import AnnotationCustomFieldDefinition
from annotation_schemas import AnnotationPriceItemInput, AnnotationProjectCreate
from crud import get_users_by_role_names
from database import SessionLocal
from models import AppUser, Client, SubClient
from schemas import ProjectRoleAssignmentInput


EXPECTED_HEADERS = [
    "项目标签（0828开始-按照【客户名-项目名-日期】登记）",
    "路径",
    "跟进状态",
    "项目状态",
    "项目部",
    "客户部",
    "客户名称",
    "子客户/对接人",
    "客户来源",
    "咨询时间",
    "开始合作时间",
    "项目类型",
    "客户（潜在）需求量",
    "价格（单价）",
    "信实汇报成交量",
    "信实实际承接量",
    "产出数据",
    "客户验收数据",
    "总价（结算）",
]

PLACEHOLDERS = {"", "无", "暂无", "无资料", "none", "null", "-", "—"}

STATUS_MAP = {
    "已取消": "cancelled",
    "已部分取消": "partially_cancelled",
    "进行中": "project_in_progress",
    "试标中": "trial_in_progress",
    "试标完：跟进结果": "client_feedback",
    "待确认": "initial_consultation",
    "已结束": "client_feedback",
    "已验收开票": "client_feedback",
    "已付款": "client_feedback",
    "暂结束": "partially_cancelled",
}

USER_ALIASES = {
    "伟豪": "陈伟豪",
    "胜辉": "李胜辉",
    "景瀚": "肖景瀚",
    "静玲": "陈静玲",
}

CLIENT_ALIASES = {
    "北京数据堂": "数据堂",
    "北京海天瑞声科技": "海天瑞声",
    "刘先生（个人）": "刘先生",
    "核数聚": "苏州核数聚信息科技",
}

RESERVED_COLUMNS = [
    "项目状态",
    "客户来源",
    "价格（单价）",
    "信实汇报成交量",
    "信实实际承接量",
    "产出数据",
    "客户验收数据",
    "总价（结算）",
]


@dataclass
class SourceRow:
    row_number: int
    values: dict[str, Any]


def clean_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text_value = str(value).strip()
    return None if text_value.casefold() in PLACEHOLDERS else text_value


def normalize_key(value: Any) -> str:
    text_value = unicodedata.normalize("NFKC", clean_text(value) or "").casefold()
    return "".join(re.findall(r"[0-9a-z\u4e00-\u9fff]+", text_value))


def normalize_path(value: Any) -> str:
    text_value = clean_text(value)
    return text_value.replace("/", "\\").rstrip("\\").casefold() if text_value else ""


def parse_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text_value = clean_text(value)
    if not text_value:
        return None
    normalized = text_value.replace("年", ".").replace("月", ".").replace("日", "")
    normalized = normalized.replace("/", ".").replace("-", ".")
    match = re.fullmatch(r"\s*(20\d{2})\.(\d{1,2})\.(\d{1,2})\s*", normalized)
    if not match:
        return None
    try:
        return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def extract_project_date(project_name: Any, consultation_time: Any = None) -> Optional[str]:
    name = clean_text(project_name) or ""
    match = re.search(r"(?<!\d)(20\d{6})(?!\d)", name)
    if not match:
        match = re.search(r"(?<!\d)(26\d{4})(?!\d)", name)
        if match:
            return "20" + match.group(1)
    if match:
        return match.group(1)
    parsed = parse_datetime(consultation_time)
    return parsed.strftime("%Y%m%d") if parsed else None


def infer_client_name(row: SourceRow) -> Optional[str]:
    explicit = clean_text(row.values.get("客户名称"))
    if explicit:
        return explicit
    name = clean_text(row.values.get(EXPECTED_HEADERS[0])) or ""
    content = name.strip("【】[] ")
    parts = [part.strip() for part in content.split("-") if part.strip()]
    if parts and re.fullmatch(r"(?:20)?\d{6,8}(?:-\d+)?", parts[0]):
        parts = parts[1:]
    if parts:
        return parts[0]
    return None


def map_project_types(row: SourceRow) -> list[str]:
    source = " ".join(filter(None, [
        clean_text(row.values.get("项目类型")),
        clean_text(row.values.get(EXPECTED_HEADERS[0])),
    ])).casefold()
    result: list[str] = []

    def add(code: str, *signals: str) -> None:
        if any(signal.casefold() in source for signal in signals) and code not in result:
            result.append(code)

    add("audio_collection", "采集", "录音", "双人对话")
    add("audio_annotation", "音频标注", "语音标注", "asr", "转写", "字幕对齐", "转录", "ast")
    add("audio_evaluation", "音频评测", "音频测评", "语音评测", "语音测评")
    add("text_evaluation", "文本评测", "文本测评")
    add("text_annotation", "文本标注", "文字标注", "tn文字", "文本归一化", "文本项目")
    add("quality_inspection", "质检", "验收", "语料把控", "审核", "校对")
    add("listening_test", "测听")
    add("slot_deduction", "扣槽")
    add("generalization", "泛化")
    add("translation", "翻译", "互译", "转译", "翻配")
    if not result and any(signal in source for signal in ("测评", "评测", "评估训练")):
        result.append("text_evaluation")
    if not result and any(signal in source for signal in ("标注", "asr", "采标")):
        result.append("text_annotation")
    return result


def parse_simple_price(value: Any, project_types: list[str]) -> list[AnnotationPriceItemInput]:
    text_value = clean_text(value)
    if not text_value or "\n" in text_value or any(token in text_value for token in ("左右", "约", "~", "～")):
        return []
    if re.search(r"\d\s*[-至到]\s*\d", text_value):
        return []
    match = re.fullmatch(r"\s*(?:每)?\s*(\d+(?:\.\d+)?)\s*元\s*/?\s*([^，,；;]+?)\s*", text_value)
    if not match:
        return []
    unit = match.group(2).strip()
    if not unit:
        return []
    return [AnnotationPriceItemInput(
        project_type=project_types[0] if project_types else None,
        amount=Decimal(match.group(1)),
        currency="CNY",
        unit=unit,
        remarks="由旧台账单一明确价格自动整理；请人工复核",
    )]


def load_source(path: Path) -> list[SourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    headers = [clean_text(cell.value) or "" for cell in worksheet[1]]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise ValueError("源表表头与预期不一致，请先人工确认文件版本")
    result = []
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        values = dict(zip(headers, cells))
        if any(clean_text(value) for value in cells):
            result.append(SourceRow(row_number=row_number, values=values))
    return result


def user_name_map(users: list[AppUser]) -> dict[str, AppUser]:
    result = {}
    for user in users:
        for value in (user.full_name, user.username):
            key = normalize_key(value)
            if key:
                result.setdefault(key, user)
    return result


def active_user_map(db) -> dict[str, AppUser]:
    return user_name_map(db.query(AppUser).filter(AppUser.is_active.is_(True)).all())


def resolve_single_user(raw_value: Any, users: dict[str, AppUser]) -> Optional[AppUser]:
    text_value = clean_text(raw_value)
    if not text_value or re.search(r"[\n、,，/&]", text_value):
        return None
    aliased = USER_ALIASES.get(text_value.strip(), text_value.strip())
    return users.get(normalize_key(aliased))


def client_index(db) -> tuple[dict[str, tuple[Client, Optional[SubClient]]], list[tuple[str, Client, Optional[SubClient]]]]:
    exact: dict[str, tuple[Client, Optional[SubClient]]] = {}
    candidates: list[tuple[str, Client, Optional[SubClient]]] = []
    for client in db.query(Client).order_by(Client.created_at.asc(), Client.id.asc()).all():
        for value in (client.client_short_name, client.client_name):
            key = normalize_key(value)
            if key:
                exact.setdefault(key, (client, None))
                candidates.append((key, client, None))
    for sub_client in db.query(SubClient).order_by(SubClient.created_at.asc(), SubClient.id.asc()).all():
        for value in (sub_client.client_short_name, sub_client.client_name):
            key = normalize_key(value)
            if key:
                exact.setdefault(key, (sub_client.parent_client, sub_client))
                candidates.append((key, sub_client.parent_client, sub_client))
    return exact, candidates


def resolve_client(
    raw_name: Any,
    exact: dict[str, tuple[Client, Optional[SubClient]]],
    candidates: list[tuple[str, Client, Optional[SubClient]]],
) -> tuple[Optional[Client], Optional[SubClient], str]:
    alias = CLIENT_ALIASES.get(clean_text(raw_name) or "", clean_text(raw_name))
    key = normalize_key(alias)
    if not key:
        return None, None, "missing"
    if key in exact:
        client, sub_client = exact[key]
        return client, sub_client, "exact"
    scored = sorted(
        ((SequenceMatcher(None, key, item_key).ratio(), client, sub_client) for item_key, client, sub_client in candidates),
        key=lambda item: item[0],
        reverse=True,
    )
    if scored and scored[0][0] >= 0.92 and (len(scored) == 1 or scored[0][0] - scored[1][0] >= 0.05):
        return scored[0][1], scored[0][2], f"fuzzy:{scored[0][0]:.3f}"
    return None, None, "new"


def build_description(row: SourceRow, parsed_price: bool, extra_reserved: dict[str, Any]) -> Optional[str]:
    lines = [f"旧台账来源行：{row.row_number}"]
    raw_type = clean_text(row.values.get("项目类型"))
    if raw_type:
        lines.append(f"原项目类型：{raw_type}")
    reserved = []
    for column in RESERVED_COLUMNS:
        value = clean_text(row.values.get(column))
        if value:
            suffix = "（已同时整理为价格明细）" if column == "价格（单价）" and parsed_price else ""
            reserved.append(f"{column}：{value}{suffix}")
    for label, value in extra_reserved.items():
        value = clean_text(value)
        if value:
            reserved.append(f"{label}：{value}")
    if reserved:
        lines.extend(["", "【待人工处理（原表预留）】", *reserved])
    return "\n".join(lines) or None


def match_existing(row: SourceRow, client_name: Optional[str], existing: list[AnnotationProject]) -> tuple[Optional[AnnotationProject], str]:
    project_name = clean_text(row.values.get(EXPECTED_HEADERS[0])) or ""
    name_key = normalize_key(project_name)
    path_key = normalize_path(row.values.get("路径"))
    source_date = extract_project_date(project_name, row.values.get("咨询时间"))
    client_key = normalize_key(client_name)
    for project in existing:
        if name_key and name_key == normalize_key(project.project_name):
            return project, "same_name"
        if path_key and path_key == normalize_path(project.project_path):
            return project, "same_path"
    best: tuple[float, Optional[AnnotationProject]] = (0.0, None)
    if name_key and source_date and client_key:
        for project in existing:
            if client_key != normalize_key(project.client_full_name):
                continue
            if source_date != extract_project_date(project.project_name, project.customer_consultation_time):
                continue
            score = SequenceMatcher(None, name_key, normalize_key(project.project_name)).ratio()
            if score > best[0]:
                best = (score, project)
    if best[1] is not None and best[0] >= 0.78:
        return best[1], f"similar_name:{best[0]:.3f}"
    return None, ""


def safe_project_path(value: Any, extra_reserved: dict[str, Any]) -> Optional[str]:
    text_value = clean_text(value)
    if not text_value:
        return None
    try:
        from path_security import validate_managed_path
        return validate_managed_path(text_value)
    except ValueError:
        extra_reserved["原项目路径（未通过路径规则）"] = text_value
        return None


def prepare_payload(
    row: SourceRow,
    db,
    users: dict[str, AppUser],
    project_manager_users: dict[str, AppUser],
    custom_fields: dict[str, AnnotationCustomFieldDefinition],
    clients_exact,
    client_candidates,
) -> tuple[AnnotationProjectCreate, dict[str, Any]]:
    values = row.values
    project_name = clean_text(values.get(EXPECTED_HEADERS[0]))
    raw_client_name = infer_client_name(row)
    client, sub_client, client_match = resolve_client(raw_client_name, clients_exact, client_candidates)
    customer_manager = resolve_single_user(values.get("客户部"), users)
    project_manager = resolve_single_user(values.get("项目部"), project_manager_users)
    project_types = map_project_types(row)
    prices = parse_simple_price(values.get("价格（单价）"), project_types)
    extra_reserved: dict[str, Any] = {}
    consultation_time = parse_datetime(values.get("咨询时间"))
    confirmation_time = parse_datetime(values.get("开始合作时间"))
    if clean_text(values.get("咨询时间")) and not consultation_time:
        extra_reserved["咨询时间（未识别）"] = values.get("咨询时间")
    if clean_text(values.get("开始合作时间")) and not confirmation_time:
        extra_reserved["开始合作时间（未识别）"] = values.get("开始合作时间")
    if clean_text(values.get("客户部")) and not customer_manager:
        extra_reserved["客户部（未匹配唯一用户）"] = values.get("客户部")
    if clean_text(values.get("项目部")) and not project_manager:
        extra_reserved["项目部（未匹配唯一项目经理用户）"] = values.get("项目部")
    if raw_client_name and not client:
        extra_reserved["客户名称（系统无精确匹配，将创建待完善客户）"] = raw_client_name
    project_path = safe_project_path(values.get("路径"), extra_reserved)
    custom_values = {}
    if clean_text(values.get("跟进状态")) and "跟进状态" in custom_fields:
        custom_values[str(custom_fields["跟进状态"].id)] = clean_text(values.get("跟进状态"))
    payload = AnnotationProjectCreate(
        project_name=project_name,
        project_types=project_types,
        task_description=build_description(row, bool(prices), extra_reserved),
        client_id=client.id if client else None,
        sub_client_id=sub_client.id if sub_client else None,
        client_name=raw_client_name if not client else None,
        client_short_name=raw_client_name if not client else None,
        contact_name=clean_text(values.get("子客户/对接人")),
        project_status=STATUS_MAP.get(clean_text(values.get("项目状态")) or "", "initial_consultation"),
        status_effective_on=date.today(),
        custom_values=custom_values,
        potential_demand=clean_text(values.get("客户（潜在）需求量")),
        project_path=project_path,
        customer_consultation_time=consultation_time,
        customer_confirmation_time=confirmation_time,
        client_manager_id=customer_manager.id if customer_manager else None,
        role_assignments=[
            ProjectRoleAssignmentInput(
                role_code="project_manager",
                assignee_id=project_manager.id,
            )
        ] if project_manager else [],
        price_items=prices,
    )
    meta = {
        "row": row.row_number,
        "project_name": project_name,
        "source_client": raw_client_name,
        "client_match": client_match,
        "resolved_client": client.client_name if client else None,
        "project_types": project_types,
        "project_status": payload.project_status,
        "price_items": len(prices),
    }
    return payload, meta


def run(args: argparse.Namespace) -> dict[str, Any]:
    source_path = Path(args.source).resolve()
    source_hash = hashlib.sha256(source_path.read_bytes()).hexdigest()
    rows = load_source(source_path)
    only_rows = {int(value) for value in args.only_rows.split(",") if value.strip()}
    force_rows = {int(value) for value in args.force_rows.split(",") if value.strip()}
    selected_rows = [row for row in rows if not only_rows or row.row_number in only_rows]
    db = SessionLocal()
    report: dict[str, Any] = {
        "source": str(source_path),
        "source_sha256": source_hash,
        "mode": "apply" if args.apply else "dry-run",
        "host": socket.gethostname(),
        "total_source_rows": len(rows),
        "selected_source_rows": len(selected_rows),
        "created": [],
        "skipped": [],
        "failed": [],
        "prepared": [],
    }
    try:
        if args.apply and socket.gethostname().casefold() != args.expected_host.casefold():
            raise RuntimeError(f"主机校验失败：当前为 {socket.gethostname()}，预期为 {args.expected_host}")
        users = active_user_map(db)
        project_manager_users = user_name_map(get_users_by_role_names(db, ["项目经理"]))
        clients_exact, client_candidates = client_index(db)
        custom_fields = {
            item.field_label: item
            for item in db.query(AnnotationCustomFieldDefinition).filter(
                AnnotationCustomFieldDefinition.table_code == "project",
                AnnotationCustomFieldDefinition.project_id.is_(None),
                AnnotationCustomFieldDefinition.is_active.is_(True),
            ).all()
        }
        existing = db.query(AnnotationProject).order_by(AnnotationProject.created_at.asc()).all()
        creator = db.query(AppUser).filter(
            AppUser.username == args.created_by,
            AppUser.is_active.is_(True),
        ).first()
        if args.apply and not creator:
            raise RuntimeError(f"未找到启用中的导入用户：{args.created_by}")
        for row in selected_rows:
            idempotency_key = f"legacy-annotation-{source_hash[:16]}-row-{row.row_number}"
            known = db.query(AnnotationProject).filter(
                AnnotationProject.idempotency_key == idempotency_key
            ).first()
            client_name = infer_client_name(row)
            duplicate, duplicate_reason = (
                (None, "") if row.row_number in force_rows
                else match_existing(row, client_name, existing)
            )
            if known or duplicate:
                matched = known or duplicate
                report["skipped"].append({
                    "row": row.row_number,
                    "project_name": clean_text(row.values.get(EXPECTED_HEADERS[0])),
                    "reason": "same_import_key" if known else duplicate_reason,
                    "existing_id": str(matched.id),
                    "existing_order_no": matched.order_no,
                    "existing_name": matched.project_name,
                })
                continue
            try:
                payload, meta = prepare_payload(
                    row, db, users, project_manager_users, custom_fields, clients_exact, client_candidates,
                )
                if not args.apply:
                    report["prepared"].append(meta)
                    continue
                project = annotation_service.create_annotation_project(
                    db,
                    payload,
                    creator.id,
                    idempotency_key=idempotency_key,
                )
                report["created"].append({
                    **meta,
                    "id": str(project.id),
                    "order_no": project.order_no,
                    "client_id": str(project.client_id) if project.client_id else None,
                    "client_created_by_import": meta["client_match"] == "new",
                })
            except Exception as exc:  # 单行失败不阻断其余可导入数据
                db.rollback()
                report["failed"].append({
                    "row": row.row_number,
                    "project_name": clean_text(row.values.get(EXPECTED_HEADERS[0])),
                    "error": str(exc),
                })
        report["summary"] = {
            "created": len(report["created"]),
            "prepared": len(report["prepared"]),
            "skipped": len(report["skipped"]),
            "failed": len(report["failed"]),
        }
        return report
    finally:
        db.close()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", help="旧标注项目 XLSX 文件路径")
    parser.add_argument("--report", required=True, help="JSON 报告输出路径")
    parser.add_argument("--apply", action="store_true", help="实际写入数据库；默认仅预览")
    parser.add_argument("--expected-host", default="WIN-LOLJ8UHT2G5", help="写入时允许的主机名")
    parser.add_argument("--created-by", default="admin", help="导入记录创建人用户名")
    parser.add_argument("--only-rows", default="", help="仅处理指定 Excel 行号，逗号分隔")
    parser.add_argument("--force-rows", default="", help="指定行忽略名称/路径近似去重，仍保留幂等键校验")
    args = parser.parse_args()
    report = run(args)
    report_path = Path(args.report).resolve()
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
