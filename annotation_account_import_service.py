"""标注项目账号 XLSX 预览与导入。"""

from __future__ import annotations

import json
import re
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from annotation_custom_field_service import create_custom_field, list_custom_fields
from annotation_models import AnnotationProject, AnnotationProjectLanguageItem
from annotation_ops_models import (
    AnnotationCustomFieldDefinition,
    AnnotationPlatform,
    AnnotationPlatformAccount,
)
from annotation_ops_schemas import AccountBatchRow, AccountWrite, CustomFieldWrite
from annotation_ops_service import batch_save_accounts
from resource_models import ResourceAnnotationLanguageSkill, ResourceCapability, ResourcePerson


MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_ROWS = 5000
HEADER_SCAN_ROWS = 20

FIXED_TARGETS = {
    "login_account", "password", "nickname", "person_name", "gender", "ignore",
}
CUSTOM_DATA_TYPES = {"text", "number", "date", "datetime", "boolean", "single_select", "multi_select", "url"}


def _normalized_header(value: Any) -> str:
    return re.sub(r"[\s_\-（）()]+", "", str(value or "").strip().casefold())


def _safe_field_key(label: str, fallback: str) -> str:
    aliases = {
        "数据编号": "external_data_no",
        "状态栏": "quality_status",
        "质检状态": "quality_status",
        "价格": "price",
        "错误点": "error_feedback",
        "问题": "issue_feedback",
        "错误点/问题": "error_feedback",
        "标红需反馈": "highlight_feedback",
    }
    if label in aliases:
        return aliases[label]
    ascii_key = re.sub(r"[^a-z0-9]+", "_", label.casefold()).strip("_")
    if not ascii_key or not ascii_key[0].isalpha():
        ascii_key = fallback
    return ascii_key[:100]


def _suggest_target(label: str, duplicate_index: int) -> dict[str, Any]:
    normalized = _normalized_header(label)
    if normalized in {"useremail", "email", "登录账号", "登录邮箱", "账号"}:
        return {"target": "login_account"}
    if normalized in {"initialpassword", "password", "初始密码", "密码"}:
        return {"target": "password"}
    if normalized in {"所分配人员姓名", "分配人员", "标注员", "姓名"}:
        return {"target": "person_name"}
    if normalized in {"性别", "gender"}:
        return {"target": "gender"}
    if normalized in {"账号昵称", "昵称", "nickname"}:
        return {"target": "nickname"}
    data_type = "number" if normalized in {"价格", "单价", "price"} else "text"
    suffix = f"_{duplicate_index}" if duplicate_index > 1 else ""
    key = f"{_safe_field_key(label, 'import_field')}{suffix}"
    return {
        "target": "new_custom",
        "fieldKey": key,
        "fieldLabel": f"{label}{f'（{duplicate_index}）' if duplicate_index > 1 else ''}",
        "dataType": data_type,
    }


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return value


def _has_value(value: Any) -> bool:
    return value is not None and value != "" and value != []


def _load_workbook(content: bytes):
    if not content:
        raise ValueError("请选择 XLSX 文件")
    if len(content) > MAX_FILE_SIZE:
        raise ValueError("文件不能超过 10MB")
    try:
        return load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("文件无法解析，请确认是有效的 .xlsx 文件") from exc


def _known_header_score(values: list[Any]) -> int:
    known = {
        "数据编号", "useremail", "email", "initialpassword", "password", "所分配人员姓名",
        "标注员", "性别", "gender", "状态栏", "质检状态", "价格", "错误点/问题", "标红需反馈",
    }
    return sum(1 for value in values if _normalized_header(value) in known)


def _detect_header_row(sheet) -> int:
    best_row, best_score = 1, -1
    for row_no, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, HEADER_SCAN_ROWS), values_only=True), start=1):
        score = _known_header_score(list(row))
        if score > best_score:
            best_row, best_score = row_no, score
    return best_row


def _headers(sheet, header_row: int) -> list[dict[str, Any]]:
    values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    occurrences: dict[str, int] = {}
    result = []
    for index, raw in enumerate(values):
        label = str(raw or "").strip()
        if not label:
            continue
        occurrences[label] = occurrences.get(label, 0) + 1
        duplicate_index = occurrences[label]
        unique_label = label if duplicate_index == 1 else f"{label}（第{duplicate_index}列）"
        result.append({
            "index": index,
            "label": label,
            "uniqueLabel": unique_label,
            "suggestion": _suggest_target(label, duplicate_index),
        })
    return result


def _normalize_mapping(headers: list[dict[str, Any]], mapping: Any) -> list[dict[str, Any]]:
    supplied = mapping if isinstance(mapping, list) else []
    by_index = {int(item["index"]): item for item in supplied if isinstance(item, dict) and "index" in item}
    result = []
    used_keys: set[str] = set()
    for header in headers:
        item = {"index": header["index"], **(by_index.get(header["index"]) or header["suggestion"])}
        target = item.get("target", "ignore")
        if target not in FIXED_TARGETS | {"custom", "new_custom"}:
            raise ValueError(f"字段“{header['uniqueLabel']}”映射目标无效")
        if target == "new_custom":
            key = str(item.get("fieldKey") or "").strip()
            if not re.fullmatch(r"[a-z][a-z0-9_]{0,99}", key):
                raise ValueError(f"字段“{header['uniqueLabel']}”的字段键无效")
            original = key
            suffix = 2
            while key in used_keys:
                key = f"{original[:95]}_{suffix}"
                suffix += 1
            item["fieldKey"] = key
            used_keys.add(key)
            data_type = str(item.get("dataType") or "text")
            if data_type not in CUSTOM_DATA_TYPES:
                raise ValueError(f"字段“{header['uniqueLabel']}”的数据类型无效")
            if data_type in {"single_select", "multi_select"} and not item.get("options"):
                raise ValueError(f"字段“{header['uniqueLabel']}”的选择项不能为空")
        result.append(item)
    return result


def _parse_rows(sheet, header_row: int, headers: list[dict[str, Any]], mapping: list[dict[str, Any]]) -> list[dict[str, Any]]:
    index_to_header = {item["index"]: item for item in headers}
    rows = []
    for row_no, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        projected = {index: _cell_value(values[index]) if index < len(values) else None for index in index_to_header}
        if not any(_has_value(value) for value in projected.values()):
            continue
        if len(rows) >= MAX_ROWS:
            raise ValueError("有效数据不能超过 5000 行")
        fixed: dict[str, Any] = {}
        custom: dict[str, Any] = {}
        display = {}
        for rule in mapping:
            value = projected.get(rule["index"])
            display[index_to_header[rule["index"]]["uniqueLabel"]] = value
            target = rule.get("target")
            if target in FIXED_TARGETS - {"ignore"}:
                fixed[target] = value
            elif target in {"custom", "new_custom"}:
                custom[str(rule.get("fieldId") or rule.get("fieldKey"))] = value
        rows.append({"rowNumber": row_no, "fixed": fixed, "custom": custom, "values": display})
    return rows


def _validate_defaults(db: Session, defaults: dict[str, Any]) -> tuple[UUID, UUID, UUID, list[UUID]]:
    try:
        client_id = UUID(str(defaults["clientId"]))
        project_id = UUID(str(defaults["projectId"]))
        platform_id = UUID(str(defaults["platformId"]))
        language_ids = [UUID(str(value)) for value in defaults.get("languageItemIds") or []]
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("客户、项目、平台和语言方向为必填项") from exc
    platform = db.get(AnnotationPlatform, platform_id)
    project = db.get(AnnotationProject, project_id)
    if not platform or platform.client_id != client_id:
        raise ValueError("所选平台不属于当前客户")
    if not project or project.client_id != client_id:
        raise ValueError("所选项目不属于当前客户")
    if not language_ids:
        raise ValueError("至少选择一个语言方向")
    found = {row.id for row in db.query(AnnotationProjectLanguageItem).filter(
        AnnotationProjectLanguageItem.project_id == project_id,
        AnnotationProjectLanguageItem.id.in_(language_ids),
    ).all()}
    if found != set(language_ids):
        raise ValueError("语言方向不属于当前项目")
    return client_id, project_id, platform_id, language_ids


def _person_matches(db: Session, name: str) -> list[ResourcePerson]:
    return db.query(ResourcePerson).filter(func.lower(ResourcePerson.full_name) == name.strip().lower()).all()


def _preview_actions(db: Session, rows: list[dict[str, Any]], platform_id: UUID) -> list[dict[str, Any]]:
    seen: set[str] = set()
    result = []
    for row in rows:
        fixed = row["fixed"]
        login = str(fixed.get("login_account") or "").strip()
        password = str(fixed.get("password") or "").strip()
        person_name = str(fixed.get("person_name") or "").strip()
        gender = str(fixed.get("gender") or "").strip()
        errors, warnings = [], []
        existing = None
        normalized = login.casefold()
        if not login:
            errors.append("登录账号不能为空")
        elif normalized in seen:
            errors.append("文件内登录账号重复，后续行不会导入")
        else:
            seen.add(normalized)
            existing = db.query(AnnotationPlatformAccount).filter(
                AnnotationPlatformAccount.platform_id == platform_id,
                AnnotationPlatformAccount.login_account_normalized == normalized,
            ).first()
            if existing:
                warnings.append("将更新已存在账号的非空字段")
            elif not password:
                errors.append("新增账号必须提供密码")
        person_action = None
        if person_name:
            people = _person_matches(db, person_name)
            if len(people) > 1:
                errors.append("人员姓名匹配不唯一")
            elif people:
                person_action = "match"
                if gender and people[0].gender and people[0].gender != gender:
                    warnings.append(f"人才档案性别为“{people[0].gender}”，不会被“{gender}”覆盖")
                elif gender and not people[0].gender:
                    warnings.append("将补充人才档案性别")
            else:
                person_action = "create"
                warnings.append("将创建待完善标注员档案")
        action = "error" if errors else ("update" if existing else "create")
        result.append({
            "rowNumber": row["rowNumber"], "values": row["values"], "action": action,
            "personAction": person_action, "warnings": warnings, "errors": errors,
        })
    return result


def parse_import_payload(
    db: Session, content: bytes, defaults: dict[str, Any], *, sheet_name: str | None = None,
    header_row: int | None = None, mapping: Any = None,
) -> dict[str, Any]:
    workbook = _load_workbook(content)
    if sheet_name and sheet_name not in workbook.sheetnames:
        raise ValueError("所选工作表不存在")
    selected_name = sheet_name or workbook.sheetnames[0]
    sheet = workbook[selected_name]
    selected_header_row = header_row or _detect_header_row(sheet)
    headers = _headers(sheet, selected_header_row)
    normalized_mapping = _normalize_mapping(headers, mapping)
    rows = _parse_rows(sheet, selected_header_row, headers, normalized_mapping)
    _, project_id, platform_id, _ = _validate_defaults(db, defaults)
    existing_fields = list_custom_fields(db, "account_assignment", project_id, include_inactive=True)
    previews = _preview_actions(db, rows, platform_id)
    required_fields = [item for item in existing_fields if item.is_active and item.is_required]
    mapped_required = {
        str(rule.get("fieldId")) for rule in normalized_mapping
        if rule.get("target") == "custom" and rule.get("fieldId")
    }
    for source_row, preview in zip(rows, previews):
        for field in required_fields:
            field_id = str(field.id)
            if field_id not in mapped_required:
                preview["errors"].append(f"必填项目字段“{field.field_label}”尚未映射")
            elif not _has_value(source_row["custom"].get(field_id)):
                preview["errors"].append(f"必填项目字段“{field.field_label}”不能为空")
        if preview["errors"]:
            preview["action"] = "error"
    return {
        "sheets": workbook.sheetnames, "sheetName": selected_name, "headerRow": selected_header_row,
        "headers": headers, "mapping": normalized_mapping,
        "projectFields": [
            {"id": str(item.id), "fieldKey": item.field_key, "fieldLabel": item.field_label, "dataType": item.data_type}
            for item in existing_fields if item.is_active
        ],
        "rows": previews,
        "summary": {
            "total": len(previews),
            "create": sum(item["action"] == "create" for item in previews),
            "update": sum(item["action"] == "update" for item in previews),
            "error": sum(bool(item["errors"]) for item in previews),
            "warning": sum(bool(item["warnings"]) for item in previews),
        },
        "_parsedRows": rows,
    }


def _ensure_custom_fields(db: Session, project_id: UUID, mapping: list[dict[str, Any]], user_id: UUID | None) -> dict[str, UUID]:
    existing_rows = list_custom_fields(db, "account_assignment", project_id, include_inactive=True)
    by_key = {item.field_key: item for item in existing_rows}
    result = {item.field_key: item.id for item in existing_rows}
    for rule in mapping:
        if rule.get("target") == "custom" and rule.get("fieldId"):
            result[str(rule["fieldId"])] = UUID(str(rule["fieldId"]))
        if rule.get("target") != "new_custom":
            continue
        key = rule["fieldKey"]
        if key in result:
            existing = by_key[key]
            if existing.data_type != str(rule.get("dataType") or "text"):
                raise ValueError(f"项目字段“{key}”已存在且类型不同")
            if not existing.is_active:
                existing.is_active = True
                db.commit()
            continue
        created = create_custom_field(db, CustomFieldWrite(
            project_id=project_id, table_code="account_assignment", field_key=key,
            field_label=str(rule.get("fieldLabel") or key), data_type=str(rule.get("dataType") or "text"),
            options=rule.get("options") or [], is_required=False, is_active=True,
        ), user_id)
        result[key] = created.id
    return result


def _resolve_person(
    db: Session, name: str, gender: str, language_items: list[AnnotationProjectLanguageItem],
) -> tuple[UUID | None, list[str]]:
    if not name:
        return None, []
    people = _person_matches(db, name)
    if len(people) > 1:
        raise ValueError("人员姓名匹配不唯一")
    warnings: list[str] = []
    if people:
        person = people[0]
        if gender and not person.gender:
            person.gender = gender
        elif gender and person.gender != gender:
            warnings.append(f"保留人才档案原性别“{person.gender}”")
        capability = db.query(ResourceCapability).filter(
            ResourceCapability.person_id == person.id,
            ResourceCapability.capability_type == "annotation",
        ).first()
        if not capability:
            db.add(ResourceCapability(person_id=person.id, capability_type="annotation", status="active", source="import"))
        elif capability.status != "active":
            capability.status = "active"
    else:
        person = ResourcePerson(full_name=name, gender=gender or None, status="standby")
        db.add(person)
        db.flush()
        db.add(ResourceCapability(person_id=person.id, capability_type="annotation", status="active", source="import"))
    existing_skills = {
        (item.source_language_id, item.target_language_id)
        for item in db.query(ResourceAnnotationLanguageSkill).filter(
            ResourceAnnotationLanguageSkill.person_id == person.id
        ).all()
    }
    for item in language_items:
        key = (item.source_language_id, item.target_language_id)
        if key not in existing_skills:
            db.add(ResourceAnnotationLanguageSkill(
                person_id=person.id, source_language_id=item.source_language_id,
                target_language_id=item.target_language_id,
            ))
            existing_skills.add(key)
    db.flush()
    return person.id, warnings


def import_accounts(
    db: Session, content: bytes, defaults: dict[str, Any], user_id: UUID | None, *,
    sheet_name: str | None = None, header_row: int | None = None, mapping: Any = None,
) -> dict[str, Any]:
    parsed = parse_import_payload(
        db, content, defaults, sheet_name=sheet_name, header_row=header_row, mapping=mapping,
    )
    client_id, project_id, platform_id, language_ids = _validate_defaults(db, defaults)
    if parsed["summary"]["error"]:
        raise ValueError("预览仍有错误，请修正映射或源数据后再导入")
    field_ids = _ensure_custom_fields(db, project_id, parsed["mapping"], user_id)
    language_items = db.query(AnnotationProjectLanguageItem).filter(
        AnnotationProjectLanguageItem.id.in_(language_ids)
    ).all()
    source = str(defaults.get("accountSource") or "client_provided")
    owner_id = UUID(str(defaults["ownerId"])) if defaults.get("ownerId") else user_id
    results = []
    seen: set[str] = set()
    for source_row in parsed["_parsedRows"]:
        fixed = source_row["fixed"]
        login = str(fixed.get("login_account") or "").strip()
        normalized = login.casefold()
        if not login or normalized in seen:
            results.append({"rowNumber": source_row["rowNumber"], "success": False, "error": "登录账号为空或重复"})
            continue
        seen.add(normalized)
        try:
            existing = db.query(AnnotationPlatformAccount).filter(
                AnnotationPlatformAccount.platform_id == platform_id,
                AnnotationPlatformAccount.login_account_normalized == normalized,
            ).first()
            person_id, warnings = _resolve_person(
                db, str(fixed.get("person_name") or "").strip(),
                str(fixed.get("gender") or "").strip(), language_items,
            )
            custom_values = {}
            for rule in parsed["mapping"]:
                if rule.get("target") not in {"custom", "new_custom"}:
                    continue
                raw_key = str(rule.get("fieldId") or rule.get("fieldKey"))
                field_id = field_ids.get(raw_key)
                value = source_row["custom"].get(raw_key)
                if field_id and _has_value(value):
                    custom_values[str(field_id)] = value
            password = str(fixed.get("password") or "") or None
            nickname = str(fixed.get("nickname") or "").strip() or (existing.nickname if existing else None)
            account = AccountWrite(
                platform_id=platform_id, owner_id=owner_id, nickname=nickname,
                login_account=login, password=password,
                account_status="assigned" if person_id else "available",
                registration_status="registered" if (password or (existing and existing.password)) else "unregistered",
                account_source=source,
                expires_on=existing.expires_on if existing else None,
                remarks=existing.remarks if existing else None,
                custom_values=existing.custom_values if existing else {},
            )
            row = AccountBatchRow(
                row_key=str(source_row["rowNumber"]), id=existing.id if existing else None,
                account=account, person_id=person_id, project_id=project_id,
                language_item_ids=language_ids, assignment_custom_values=custom_values,
            )
            outcome = batch_save_accounts(db, client_id, [row], user_id)["results"][0]
            results.append({
                "rowNumber": source_row["rowNumber"], "success": outcome["success"],
                "action": "update" if existing else "create", "warnings": warnings,
                "error": outcome.get("error"),
            })
        except Exception as exc:
            db.rollback()
            results.append({"rowNumber": source_row["rowNumber"], "success": False, "error": str(exc)})
    return {
        "results": results,
        "summary": {
            "total": len(results), "success": sum(item["success"] for item in results),
            "created": sum(item["success"] and item.get("action") == "create" for item in results),
            "updated": sum(item["success"] and item.get("action") == "update" for item in results),
            "failed": sum(not item["success"] for item in results),
        },
    }


def decode_json_object(value: str, label: str) -> dict[str, Any]:
    try:
        result = json.loads(value or "{}")
    except json.JSONDecodeError as exc:
        raise ValueError(f"{label}格式无效") from exc
    if not isinstance(result, dict):
        raise ValueError(f"{label}必须是对象")
    return result


def decode_json_list(value: str | None, label: str) -> list[Any] | None:
    if not value:
        return None
    try:
        result = json.loads(value)
    except json.JSONDecodeError as exc:
        raise ValueError(f"{label}格式无效") from exc
    if not isinstance(result, list):
        raise ValueError(f"{label}必须是数组")
    return result
