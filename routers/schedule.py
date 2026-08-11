"""
排班管理 API 路由
支持按日期 CRUD 操作，项目经理每日微调排班数据
"""
from datetime import date, datetime, time, timedelta
from io import BytesIO
import json
from typing import Optional, List
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy import or_
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from database import get_db
from department_utils import department_filter_values, normalize_department
from leave_service import business_now
from models import (
    WorkSchedule, AppUser, Translator, TranslatorSchedule,
    EmployeeShiftTemplate, EmployeeShiftOverride, EmployeeShiftOverrideAudit, EmployeeShiftLock, EmployeeLeave,
)
from schemas import (
    WorkScheduleCreate, WorkScheduleUpdate, WorkScheduleResponse,
    TranslatorScheduleCreate, TranslatorScheduleUpdate, TranslatorScheduleResponse,
    EmployeeShiftTemplateUpdate, EmployeeShiftOverrideBatchUpdate, EmployeeShiftLockUpdate,
)
from routers.auth import get_current_user, require_module_access

router = APIRouter(prefix="/schedules", tags=["schedules"], dependencies=[Depends(require_module_access("schedule:read", "schedule:write"))])
workbench_router = APIRouter(prefix="/schedules", tags=["schedules"])

SHIFT_PRESETS = {
    "early_early": {"label": "早早班", "start_time": "08:30", "end_time": "18:00"},
    "early": {"label": "早班", "start_time": "09:00", "end_time": "18:30"},
    "late": {"label": "晚班", "start_time": "10:30", "end_time": "20:00"},
    "late_late": {"label": "晚晚班", "start_time": "13:30", "end_time": "21:30"},
    "weekend_duty": {"label": "周末值班", "start_time": "09:30", "end_time": "18:00"},
}
SHIFT_CODES = set(SHIFT_PRESETS) | {"custom", "off", "unassigned"}


def _time_text(value: Optional[time]) -> Optional[str]:
    return value.strftime("%H:%M") if value else None


def _shift_value(shift_code: str, start_value: Optional[time] = None, end_value: Optional[time] = None) -> dict:
    preset = SHIFT_PRESETS.get(shift_code)
    if preset:
        return {
            "shift_code": shift_code,
            "shift_label": preset["label"],
            "start_time": preset["start_time"],
            "end_time": preset["end_time"],
        }
    labels = {"custom": "自定义班次", "off": "休息", "unassigned": "未安排"}
    return {
        "shift_code": shift_code,
        "shift_label": labels.get(shift_code, shift_code),
        "start_time": _time_text(start_value),
        "end_time": _time_text(end_value),
    }


def _storage_times(shift_code: str, start_value: Optional[time], end_value: Optional[time]):
    preset = SHIFT_PRESETS.get(shift_code)
    if preset:
        return time.fromisoformat(preset["start_time"]), time.fromisoformat(preset["end_time"])
    if shift_code in {"off", "unassigned"}:
        return None, None
    return start_value, end_value

WEEKDAY_HEADER_MAP = {
    "星期一": 0,
    "星期二": 1,
    "星期三": 2,
    "星期四": 3,
    "星期五": 4,
}


def _normalize_text(value) -> str:
    return str(value or "").strip()


def _normalize_person_name(value) -> str:
    return _normalize_text(value).replace(" ", "")


def _parse_excel_datetime(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = _normalize_text(value).replace(".", "/")
    for fmt in ("%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _parse_excel_date(value):
    dt = _parse_excel_datetime(value)
    if dt:
        return dt.date()
    text = _normalize_text(value).replace(".", "/")
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _cell_to_slot_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if float(value) <= 0:
            return ""
        if float(value).is_integer() and int(value) == 1:
            return "可接稿"
        return str(int(value) if float(value).is_integer() else value)
    text = _normalize_text(value)
    if not text:
        return ""
    normalized = text.lower()
    if normalized in {"1", "y", "yes", "true", "可", "可以"}:
        return "可接稿"
    if normalized in {"0", "n", "no", "false", "否"}:
        return ""
    return text


def _parse_acceptance_status(value: str) -> tuple[str, str]:
    normalized = _normalize_text(value)
    if normalized in {"2", "本周期不可接稿", "不能接稿"}:
        return "本周期不可接稿", "cycle_blocked"
    if normalized in {"1", "可接稿"}:
        return "可接稿（按日）", "day_by_day"
    return "可接稿（按日）", "day_by_day"


def _parse_schedule_day_available(value) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) == 1
    return _normalize_text(value) == "1"


def _resolve_schedule_slot(acceptance_mode: str, day_available: bool, raw_window_value: str) -> str:
    if acceptance_mode == "cycle_blocked":
        return "本周期不可接稿"
    if not day_available:
        return ""
    return raw_window_value or "可接稿"


def _parse_translator_schedule_demo_rows(content: bytes, filename: str, db: Session):
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"暂时只支持 Excel xlsx 文件导入：{exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 内容不足，至少需要表头和一行数据")

    headers = [_normalize_text(cell) for cell in rows[0]]
    name_idx = next((i for i, h in enumerate(headers) if "姓名" in h), None)
    submit_time_idx = next((i for i, h in enumerate(headers) if "提交" in h and "时间" in h), None)
    fill_date_idx = next((i for i, h in enumerate(headers) if ("填写" in h and "日期" in h) or ("本周总" in h and "日期" in h)), None)
    remarks_idx = next((i for i, h in enumerate(headers) if "备注" in h), None)
    weekday_columns = {}
    for idx, header in enumerate(headers):
        for weekday, offset in WEEKDAY_HEADER_MAP.items():
            if weekday in header:
                weekday_columns[idx] = offset
                break

    if name_idx is None or not weekday_columns:
        raise HTTPException(status_code=400, detail="未识别到姓名列或星期列，请先保持当前导出模板格式")

    translators = db.query(Translator).all()
    translator_map = {}
    for translator in translators:
        translator_map.setdefault(_normalize_person_name(translator.translator_name), translator)

    preview_items = []
    matched_translators = set()
    unmatched_names = []

    for row_index, row in enumerate(rows[1:], start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue
        raw_name = row[name_idx] if name_idx < len(row) else None
        normalized_name = _normalize_person_name(raw_name)
        if not normalized_name:
            continue

        translator = translator_map.get(normalized_name)
        if not translator:
            unmatched_names.append(_normalize_text(raw_name))
            continue

        base_date = None
        if fill_date_idx is not None and fill_date_idx < len(row):
            base_date = _parse_excel_date(row[fill_date_idx])
        if base_date is None and submit_time_idx is not None and submit_time_idx < len(row):
            submit_dt = _parse_excel_datetime(row[submit_time_idx])
            if submit_dt:
                base_date = submit_dt.date()
        if base_date is None:
            continue

        week_monday = base_date - timedelta(days=base_date.weekday())
        remarks = _normalize_text(row[remarks_idx]) if remarks_idx is not None and remarks_idx < len(row) else ""
        submitted_at = None
        if submit_time_idx is not None and submit_time_idx < len(row):
            submitted_at = _parse_excel_datetime(row[submit_time_idx])

        for col_idx, weekday_offset in weekday_columns.items():
            if col_idx >= len(row):
                continue
            slot_text = _cell_to_slot_text(row[col_idx])
            if not slot_text:
                continue

            schedule_day = week_monday + timedelta(days=weekday_offset)
            existing = (
                db.query(TranslatorSchedule)
                .filter(
                    TranslatorSchedule.translator_id == translator.id,
                    TranslatorSchedule.schedule_date == schedule_day,
                )
                .first()
            )
            preview_items.append({
                "row_no": row_index,
                "translator_id": str(translator.id),
                "translator_name": translator.translator_name,
                "schedule_date": schedule_day.isoformat(),
                "availability_status": "available",
                "available_time_slot": slot_text,
                "remarks": remarks,
                "last_confirmed_at": submitted_at.isoformat() if submitted_at else None,
                "source_type": "excel_demo",
                "source_ref": filename,
                "action": "update" if existing else "create",
                "existing": bool(existing),
                "existing_available_time_slot": existing.available_time_slot if existing else None,
            })
            matched_translators.add(translator.translator_name)

    return {
        "file_name": filename,
        "sheet_name": sheet.title,
        "headers": headers,
        "preview_items": preview_items,
        "matched_translators": len(matched_translators),
        "unmatched_names": unmatched_names,
    }


def _parse_translator_schedule_demo_rows_v2(content: bytes, filename: str, db: Session):
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"暂时只支持 Excel xlsx 文件导入：{exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 内容不足，至少需要表头和一行数据")

    translators = db.query(Translator).all()
    translator_map = {}
    for translator in translators:
        translator_map.setdefault(_normalize_person_name(translator.translator_name), translator)

    preview_items = []
    matched_translators = set()
    unmatched_names = []

    for row_index, row in enumerate(rows[1:], start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue

        raw_name = row[6] if len(row) > 6 else None
        normalized_name = _normalize_person_name(raw_name)
        if not normalized_name:
            continue

        translator = translator_map.get(normalized_name)
        if not translator:
            unmatched_names.append(_normalize_text(raw_name))
            continue

        fill_date = _parse_excel_date(row[7] if len(row) > 7 else None)
        if fill_date is None:
            continue

        submitted_at = _parse_excel_datetime(row[1] if len(row) > 1 else None)
        acceptance_raw = _normalize_text(row[8] if len(row) > 8 else None)
        acceptance_status, acceptance_mode = _parse_acceptance_status(acceptance_raw)
        raw_window_value = _normalize_text(row[9] if len(row) > 9 else None)
        raw_remark_value = _normalize_text(row[15] if len(row) > 15 else None)
        raw_payload = {
            "acceptance_raw": acceptance_raw,
            "acceptance_status": acceptance_status,
            "acceptance_mode": acceptance_mode,
            "time_slot": raw_window_value,
            "note": raw_remark_value,
        }

        for day_offset, col_idx in enumerate(range(10, 15)):
            day_available = _parse_schedule_day_available(row[col_idx] if len(row) > col_idx else None)
            slot_text = _resolve_schedule_slot(acceptance_mode, day_available, raw_window_value)
            availability_status = (
                "cycle_blocked" if acceptance_mode == "cycle_blocked"
                else ("available" if day_available else "unavailable")
            )

            schedule_day = fill_date + timedelta(days=day_offset)
            existing = (
                db.query(TranslatorSchedule)
                .filter(
                    TranslatorSchedule.translator_id == translator.id,
                    TranslatorSchedule.schedule_date == schedule_day,
                )
                .first()
            )
            preview_items.append({
                "row_no": row_index,
                "translator_id": str(translator.id),
                "translator_name": translator.translator_name,
                "fill_date": fill_date.isoformat(),
                "schedule_date": schedule_day.isoformat(),
                "availability_status": availability_status,
                "available_time_slot": slot_text,
                "acceptance_raw": acceptance_raw,
                "acceptance_status": acceptance_status,
                "acceptance_mode": acceptance_mode,
                "day_available": day_available,
                "time_slot": raw_window_value,
                "note": raw_remark_value,
                "remarks": json.dumps(raw_payload, ensure_ascii=False),
                "last_confirmed_at": submitted_at.isoformat() if submitted_at else None,
                "source_type": "excel_demo_v2",
                "source_ref": filename,
                "action": "update" if existing else "create",
                "existing": bool(existing),
                "existing_available_time_slot": existing.available_time_slot if existing else None,
            })
            matched_translators.add(translator.translator_name)

    return {
        "file_name": filename,
        "sheet_name": sheet.title,
        "headers": [_normalize_text(cell) for cell in rows[0]],
        "preview_items": preview_items,
        "matched_translators": len(matched_translators),
        "unmatched_names": unmatched_names,
    }


def _parse_translator_availability_status(value) -> str:
    normalized = _normalize_text(value).lower()
    mapping = {
        "可接稿": "available", "可接": "available", "available": "available", "1": "available",
        "不可接稿": "unavailable", "不可接": "unavailable", "unavailable": "unavailable", "0": "unavailable",
        "本周期不可接稿": "cycle_blocked", "周期不可接": "cycle_blocked", "cycle_blocked": "cycle_blocked", "2": "cycle_blocked",
    }
    status_value = mapping.get(normalized)
    if not status_value:
        raise ValueError(f"无法识别接稿状态：{value}")
    return status_value


def _parse_translator_schedule_standard_rows(content: bytes, filename: str, db: Session):
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"暂时只支持 Excel xlsx 文件导入：{exc}") from exc
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 内容不足，至少需要表头和一行数据")
    headers = [_normalize_text(cell) for cell in rows[0]]

    def find_header(*aliases):
        return next((index for index, header in enumerate(headers) if header in aliases), None)

    id_idx = find_header("译员ID", "译员 Id", "translator_id")
    name_idx = find_header("译员姓名", "姓名", "translator_name")
    date_idx = find_header("日期", "排期日期", "schedule_date")
    status_idx = find_header("接稿状态", "状态", "availability_status")
    slot_idx = find_header("可接时段", "时段", "available_time_slot")
    capacity_idx = find_header("剩余容量", "剩余字数", "remaining_capacity")
    remarks_idx = find_header("备注", "remarks")
    if date_idx is None or status_idx is None or (id_idx is None and name_idx is None):
        raise HTTPException(status_code=400, detail="标准模板必须包含译员ID或姓名、日期和接稿状态")

    translators = db.query(Translator).all()
    by_id = {str(item.id): item for item in translators}
    by_name: dict[str, list[Translator]] = {}
    for item in translators:
        by_name.setdefault(_normalize_person_name(item.translator_name), []).append(item)

    preview_items = []
    matched_translators = set()
    unmatched_names = []
    errors = []
    for row_no, row in enumerate(rows[1:], start=2):
        if not row or all(cell in (None, "") for cell in row):
            continue
        raw_id = _normalize_text(row[id_idx]) if id_idx is not None and id_idx < len(row) else ""
        raw_name = _normalize_text(row[name_idx]) if name_idx is not None and name_idx < len(row) else ""
        translator = by_id.get(raw_id) if raw_id else None
        if not translator and raw_name:
            matches = by_name.get(_normalize_person_name(raw_name), [])
            if len(matches) == 1:
                translator = matches[0]
            elif len(matches) > 1:
                errors.append({"row_no": row_no, "message": f"译员姓名重复，请填写译员ID：{raw_name}"})
                continue
        if not translator:
            unmatched_names.append(raw_name or raw_id)
            continue
        schedule_day = _parse_excel_date(row[date_idx] if date_idx < len(row) else None)
        if not schedule_day:
            errors.append({"row_no": row_no, "message": "日期格式无效"})
            continue
        try:
            availability_status = _parse_translator_availability_status(row[status_idx] if status_idx < len(row) else None)
        except ValueError as exc:
            errors.append({"row_no": row_no, "message": str(exc)})
            continue
        slot_text = _normalize_text(row[slot_idx]) if slot_idx is not None and slot_idx < len(row) else ""
        capacity_raw = row[capacity_idx] if capacity_idx is not None and capacity_idx < len(row) else None
        try:
            remaining_capacity = int(capacity_raw) if capacity_raw not in (None, "") else None
        except (TypeError, ValueError):
            errors.append({"row_no": row_no, "message": "剩余容量必须是整数"})
            continue
        remarks = _normalize_text(row[remarks_idx]) if remarks_idx is not None and remarks_idx < len(row) else ""
        existing = db.query(TranslatorSchedule).filter(
            TranslatorSchedule.translator_id == translator.id,
            TranslatorSchedule.schedule_date == schedule_day,
        ).first()
        preview_items.append({
            "row_no": row_no,
            "translator_id": str(translator.id),
            "translator_name": translator.translator_name,
            "schedule_date": schedule_day.isoformat(),
            "availability_status": availability_status,
            "available_time_slot": slot_text if availability_status == "available" else "",
            "remaining_capacity": remaining_capacity,
            "remarks": remarks,
            "last_confirmed_at": None,
            "source_type": "excel_standard",
            "source_ref": filename,
            "action": "update" if existing else "create",
            "existing": bool(existing),
            "existing_available_time_slot": existing.available_time_slot if existing else None,
        })
        matched_translators.add(translator.translator_name)
    return {
        "file_name": filename,
        "sheet_name": sheet.title,
        "headers": headers,
        "format": "standard",
        "preview_items": preview_items,
        "matched_translators": len(matched_translators),
        "unmatched_names": sorted(set(filter(None, unmatched_names))),
        "errors": errors,
    }


def _parse_translator_schedule_auto(content: bytes, filename: str, db: Session):
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        headers = [_normalize_text(cell.value) for cell in next(workbook.active.iter_rows())]
        workbook.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"暂时只支持 Excel xlsx 文件导入：{exc}") from exc
    standard_markers = {"日期", "排期日期", "schedule_date"}
    status_markers = {"接稿状态", "状态", "availability_status"}
    if standard_markers.intersection(headers) and status_markers.intersection(headers):
        return _parse_translator_schedule_standard_rows(content, filename, db)
    parsed = _parse_translator_schedule_demo_rows_v2(content, filename, db)
    parsed["format"] = "external_g_p"
    parsed.setdefault("errors", [])
    return parsed


def _normalize_task(task):
    task = task or {}
    return {
        "category": task.get("category") or "",
        "project_name": task.get("project_name")
        or task.get("projectName")
        or task.get("content")
        or task.get("project_or_task")
        or task.get("projectOrTask")
        or "",
        "order_no": task.get("order_no")
        or task.get("orderNo")
        or task.get("project_no")
        or task.get("projectNo")
        or "",
        "customer_deadline_time": task.get("customer_deadline_time")
        or task.get("customerDeadlineTime")
        or task.get("deadline")
        or "",
        "project_status": task.get("project_status")
        or task.get("projectStatus")
        or task.get("status")
        or "",
    }


def _format_cloud_revision(can_cloud_edit, can_revision) -> str:
    values = []
    for value in (can_cloud_edit, can_revision):
        if value is True:
            values.append("可")
        elif value is False:
            values.append("否")
        else:
            values.append("")
    return "/".join(values).strip("/")


def _format_daily_rate(daily_accept_count, hourly_speed, daily_word_capacity) -> str:
    values = [daily_accept_count, hourly_speed, daily_word_capacity]
    if all(value in (None, "") for value in values):
        return ""
    return "/".join("" if value in (None, "") else str(value) for value in values)


def _normalize_dept_person(person):
    person = person or {}
    tasks = person.get("tasks")
    fixed_tasks = person.get("fixed_tasks") if person.get("fixed_tasks") is not None else person.get("fixedTasks")
    return {
        "name": person.get("name") or "",
        "dept": normalize_department(person.get("dept")) or "",
        "status": person.get("status") or "scheduled",
        "tasks": [_normalize_task(t) for t in tasks] if isinstance(tasks, list) else [],
        "fixed_tasks": fixed_tasks if isinstance(fixed_tasks, list) else [],
    }


def _normalize_not_scheduled(item):
    item = item or {}
    return {
        "person_name": item.get("person_name") or item.get("personName") or "",
        "department": normalize_department(item.get("department")) or "",
        "project_name": item.get("project_name")
        or item.get("projectName")
        or item.get("project_or_task")
        or item.get("projectOrTask")
        or "",
        "order_no": item.get("order_no")
        or item.get("orderNo")
        or item.get("project_no")
        or item.get("projectNo")
        or "",
        "customer_deadline_time": item.get("customer_deadline_time")
        or item.get("customerDeadlineTime")
        or item.get("deadline")
        or "",
        "remarks": item.get("remarks") or "",
    }


def _normalize_schedule_payload(payload: dict) -> dict:
    if "dept_person_data" in payload and isinstance(payload["dept_person_data"], list):
        payload["dept_person_data"] = [_normalize_dept_person(p) for p in payload["dept_person_data"]]
    if "not_scheduled_tasks" in payload and isinstance(payload["not_scheduled_tasks"], list):
        payload["not_scheduled_tasks"] = [_normalize_not_scheduled(i) for i in payload["not_scheduled_tasks"]]
    return payload


def _normalize_schedule_record(record: WorkSchedule) -> WorkSchedule:
    record.dept_person_data = [_normalize_dept_person(p) for p in (record.dept_person_data or [])]
    record.not_scheduled_tasks = [_normalize_not_scheduled(i) for i in (record.not_scheduled_tasks or [])]
    return record


def _default_shift_for_date(schedule_day: date) -> dict:
    code = "off" if schedule_day.isoweekday() in (6, 7) else "unassigned"
    return {**_shift_value(code), "source": "unassigned", "note": None}


def _load_shift_lock_rows(db: Session, user_ids: list, through_date: date) -> dict:
    if not user_ids:
        return {}
    rows = (
        db.query(EmployeeShiftLock)
        .filter(
            EmployeeShiftLock.user_id.in_(user_ids),
            EmployeeShiftLock.effective_from <= through_date,
        )
        .order_by(EmployeeShiftLock.effective_from.asc())
        .all()
    )
    result = {}
    for row in rows:
        result.setdefault(row.user_id, []).append(row)
    return result


def _lock_for_date(lock_rows: dict, user_id, schedule_day: date):
    candidates = lock_rows.get(user_id, [])
    return next((row for row in reversed(candidates) if row.effective_from <= schedule_day), None)


def _resolve_employee_shift_rows(
    db: Session,
    users: list[AppUser],
    date_from: date,
    date_to: date,
) -> list[dict]:
    user_ids = [user.id for user in users]
    if not user_ids:
        return []
    template_rows = (
        db.query(EmployeeShiftTemplate)
        .filter(
            EmployeeShiftTemplate.user_id.in_(user_ids),
            EmployeeShiftTemplate.effective_from <= date_to,
        )
        .order_by(EmployeeShiftTemplate.effective_from.asc())
        .all()
    )
    override_rows = (
        db.query(EmployeeShiftOverride)
        .filter(
            EmployeeShiftOverride.user_id.in_(user_ids),
            EmployeeShiftOverride.schedule_date >= date_from,
            EmployeeShiftOverride.schedule_date <= date_to,
        )
        .all()
    )
    lock_rows = _load_shift_lock_rows(db, user_ids, date_to)
    leave_rows = (
        db.query(EmployeeLeave)
        .filter(
            EmployeeLeave.employee_id.in_(user_ids),
            EmployeeLeave.start_date < datetime.combine(date_to + timedelta(days=1), time.min),
            EmployeeLeave.end_date > datetime.combine(date_from, time.min),
        )
        .all()
    )
    leaves_by_user = {}
    for row in leave_rows:
        leaves_by_user.setdefault(row.employee_id, []).append(row)
    templates: dict[tuple, list[EmployeeShiftTemplate]] = {}
    for row in template_rows:
        templates.setdefault((row.user_id, row.weekday), []).append(row)
    overrides = {(row.user_id, row.schedule_date): row for row in override_rows}

    dates = [date_from + timedelta(days=offset) for offset in range((date_to - date_from).days + 1)]
    result = []
    for user in users:
        days = []
        for schedule_day in dates:
            lock = _lock_for_date(lock_rows, user.id, schedule_day)
            day_start = datetime.combine(schedule_day, time.min)
            day_end = day_start + timedelta(days=1)
            leave = next((item for item in leaves_by_user.get(user.id, []) if item.start_date < day_end and item.end_date > day_start), None)
            override = overrides.get((user.id, schedule_day))
            if override:
                value = _shift_value(override.shift_code, override.start_time, override.end_time)
                value.update({"source": "override", "note": override.note})
            else:
                candidates = templates.get((user.id, schedule_day.isoweekday()), [])
                template = next(
                    (item for item in reversed(candidates) if item.effective_from <= schedule_day),
                    None,
                )
                if template:
                    value = _shift_value(template.shift_code, template.start_time, template.end_time)
                    value.update({"source": "template", "note": None})
                else:
                    value = _default_shift_for_date(schedule_day)
            days.append({
                "date": schedule_day.isoformat(),
                **value,
                "is_locked": bool(lock and lock.is_locked),
                "lock_effective_from": lock.effective_from.isoformat() if lock else None,
                "lock_reason": lock.reason if lock else None,
                "on_leave": bool(leave),
                "leave_type": leave.leave_type if leave else None,
                "leave_start": leave.start_date.isoformat() if leave else None,
                "leave_end": leave.end_date.isoformat() if leave else None,
            })
        current_lock = _lock_for_date(lock_rows, user.id, date_from)
        result.append({
            "user_id": str(user.id),
            "name": user.full_name or user.username,
            "username": user.username,
            "department": normalize_department(user.department) or "",
            "is_locked": bool(current_lock and current_lock.is_locked),
            "lock_effective_from": current_lock.effective_from.isoformat() if current_lock else None,
            "lock_reason": current_lock.reason if current_lock else None,
            "days": days,
        })
    return result


def _employee_shift_department_options(db: Session):
    rows = (
        db.query(AppUser.department)
        .filter(AppUser.is_active == True)
        .distinct()
        .all()
    )
    departments = set()
    has_unassigned = False
    for row in rows:
        try:
            raw_value = row[0]
        except (TypeError, KeyError, IndexError):
            raw_value = getattr(row, "department", row)
        value = normalize_department(str(raw_value or "")) or ""
        if value:
            departments.add(value)
        else:
            has_unassigned = True

    options = [{"value": value, "label": value} for value in sorted(departments)]
    if has_unassigned:
        options.append({"value": "__unassigned__", "label": "未分部门"})
    return options


@router.get("/employee-shifts")
def get_employee_shifts(
    date_from: date = Query(...),
    date_to: date = Query(...),
    department: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    show_all: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: AppUser = Depends(get_current_user),
):
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    if (date_to - date_from).days > 31:
        raise HTTPException(status_code=400, detail="单次最多查询 32 天班次")
    query = db.query(AppUser).filter(AppUser.is_active == True)
    effective_department = None
    current_user_only = False
    if department == "__unassigned__":
        query = query.filter(or_(AppUser.department.is_(None), AppUser.department == ""))
        effective_department = "__unassigned__"
    elif department:
        query = query.filter(AppUser.department.in_(department_filter_values(department)))
        effective_department = normalize_department(department)
    elif not show_all:
        if current_user.department:
            query = query.filter(AppUser.department.in_(department_filter_values(current_user.department)))
            effective_department = normalize_department(current_user.department)
        else:
            # 未设置部门的账号默认只读取本人，避免退化为全员查询。
            query = query.filter(AppUser.id == current_user.id)
            current_user_only = True
    if keyword and keyword.strip():
        pattern = f"%{keyword.strip()}%"
        query = query.filter(or_(AppUser.full_name.ilike(pattern), AppUser.username.ilike(pattern)))
    users = query.order_by(AppUser.department.asc().nullsfirst(), AppUser.full_name.asc(), AppUser.username.asc()).all()
    return {
        "date_from": date_from.isoformat(),
        "date_to": date_to.isoformat(),
        "show_all": show_all and not department,
        "effective_department": effective_department,
        "current_user_only": current_user_only,
        "department_options": _employee_shift_department_options(db),
        "presets": [{"code": code, **value} for code, value in SHIFT_PRESETS.items()],
        "employees": _resolve_employee_shift_rows(db, users, date_from, date_to),
    }


@router.get("/employee-shifts/templates/{user_id}")
def get_employee_shift_template(
    user_id: UUID,
    reference_date: date = Query(default_factory=date.today),
    db: Session = Depends(get_db),
):
    user = db.query(AppUser).filter(AppUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    effective_from = (
        db.query(EmployeeShiftTemplate.effective_from)
        .filter(
            EmployeeShiftTemplate.user_id == user_id,
            EmployeeShiftTemplate.effective_from <= reference_date,
        )
        .order_by(EmployeeShiftTemplate.effective_from.desc())
        .limit(1)
        .scalar()
    )
    rows = []
    if effective_from:
        rows = (
            db.query(EmployeeShiftTemplate)
            .filter(
                EmployeeShiftTemplate.user_id == user_id,
                EmployeeShiftTemplate.effective_from == effective_from,
            )
            .order_by(EmployeeShiftTemplate.weekday.asc())
            .all()
        )
    row_map = {row.weekday: row for row in rows}
    lock_rows = _load_shift_lock_rows(db, [user.id], reference_date)
    lock = _lock_for_date(lock_rows, user.id, reference_date)
    days = []
    for weekday in range(1, 8):
        row = row_map.get(weekday)
        if row:
            days.append({"weekday": weekday, **_shift_value(row.shift_code, row.start_time, row.end_time)})
        else:
            code = "off" if weekday in (6, 7) else "unassigned"
            days.append({"weekday": weekday, **_shift_value(code)})
    monday = reference_date - timedelta(days=reference_date.weekday())
    return {
        "user_id": str(user.id),
        "name": user.full_name or user.username,
        "department": normalize_department(user.department) or "",
        "effective_from": monday.isoformat(),
        "template_effective_from": effective_from.isoformat() if effective_from else None,
        "is_locked": bool(lock and lock.is_locked),
        "lock_effective_from": lock.effective_from.isoformat() if lock else None,
        "lock_reason": lock.reason if lock else None,
        "days": days,
    }


@router.put("/employee-shifts/templates/{user_id}")
def replace_employee_shift_template(
    user_id: UUID,
    payload: EmployeeShiftTemplateUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    user = db.query(AppUser).filter(AppUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    lock = _lock_for_date(
        _load_shift_lock_rows(db, [user.id], payload.effective_from),
        user.id,
        payload.effective_from,
    )
    if lock and lock.is_locked:
        raise HTTPException(status_code=409, detail="该员工的固定班次已锁定，请先创建解锁状态后再修改模板")
    current_monday = business_now().date() - timedelta(days=business_now().date().weekday())
    exact_version_exists = db.query(EmployeeShiftTemplate.id).filter(
        EmployeeShiftTemplate.user_id == user_id,
        EmployeeShiftTemplate.effective_from == payload.effective_from,
    ).first()
    if exact_version_exists and payload.effective_from < current_monday:
        raise HTTPException(status_code=409, detail="历史常规排班版本不可修改，请选择当前或未来生效周创建新版本")
    db.query(EmployeeShiftTemplate).filter(
        EmployeeShiftTemplate.user_id == user_id,
        EmployeeShiftTemplate.effective_from == payload.effective_from,
    ).delete(synchronize_session=False)
    for day in payload.days:
        start_value, end_value = _storage_times(day.shift_code, day.start_time, day.end_time)
        db.add(EmployeeShiftTemplate(
            user_id=user_id,
            weekday=day.weekday,
            effective_from=payload.effective_from,
            shift_code=day.shift_code,
            start_time=start_value,
            end_time=end_value,
            updated_by=current_user.id,
        ))
    db.commit()
    return {"message": "周班次模板已保存", "effective_from": payload.effective_from.isoformat()}


@router.put("/employee-shifts/locks/{user_id}")
def update_employee_shift_lock(
    user_id: UUID,
    payload: EmployeeShiftLockUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    user = db.query(AppUser).filter(AppUser.id == user_id, AppUser.is_active == True).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在或已停用")
    if payload.is_locked:
        template_effective_from = db.query(EmployeeShiftTemplate.effective_from).filter(
            EmployeeShiftTemplate.user_id == user_id,
            EmployeeShiftTemplate.effective_from <= payload.effective_from,
        ).order_by(EmployeeShiftTemplate.effective_from.desc()).first()
        template_count = 0
        if template_effective_from:
            template_count = db.query(EmployeeShiftTemplate.id).filter(
                EmployeeShiftTemplate.user_id == user_id,
                EmployeeShiftTemplate.effective_from == template_effective_from[0],
            ).count()
        if template_count < 7:
            raise HTTPException(status_code=409, detail="请先为该员工配置完整的七天常规排班，再进行锁定")
    record = db.query(EmployeeShiftLock).filter(
        EmployeeShiftLock.user_id == user_id,
        EmployeeShiftLock.effective_from == payload.effective_from,
    ).first()
    current_date = business_now().date()
    current_monday = current_date - timedelta(days=current_date.weekday())
    if record and payload.effective_from < current_monday:
        raise HTTPException(status_code=409, detail="历史锁定状态不可修改，请选择当前或未来生效周创建新版本")
    if record:
        record.is_locked = payload.is_locked
        record.reason = payload.reason.strip()
        record.changed_by = current_user.id
        record.changed_at = datetime.utcnow()
    else:
        record = EmployeeShiftLock(
            user_id=user_id,
            effective_from=payload.effective_from,
            is_locked=payload.is_locked,
            reason=payload.reason.strip(),
            changed_by=current_user.id,
        )
        db.add(record)
    db.commit()
    return {
        "user_id": str(user_id),
        "effective_from": payload.effective_from.isoformat(),
        "is_locked": payload.is_locked,
        "reason": payload.reason.strip(),
    }


@router.put("/employee-shifts/overrides/batch")
def upsert_employee_shift_overrides(
    payload: EmployeeShiftOverrideBatchUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    user_ids = {item.user_id for item in payload.items}
    existing_user_ids = {
        row[0] for row in db.query(AppUser.id).filter(AppUser.id.in_(user_ids)).all()
    } if user_ids else set()
    missing = user_ids - existing_user_ids
    if missing:
        raise HTTPException(status_code=404, detail="存在无效的用户ID")
    schedule_dates = {item.schedule_date for item in payload.items}
    existing_records = (
        db.query(EmployeeShiftOverride)
        .filter(
            EmployeeShiftOverride.user_id.in_(user_ids),
            EmployeeShiftOverride.schedule_date.in_(schedule_dates),
        )
        .all()
    ) if user_ids and schedule_dates else []
    record_map = {(row.user_id, row.schedule_date): row for row in existing_records}
    lock_rows = _load_shift_lock_rows(db, list(user_ids), max(schedule_dates) if schedule_dates else date.today())
    updated = 0
    cleared = 0
    for item in payload.items:
        record = record_map.get((item.user_id, item.schedule_date))
        lock = _lock_for_date(lock_rows, item.user_id, item.schedule_date)
        if item.action == "clear":
            if record:
                db.add(EmployeeShiftOverrideAudit(
                    user_id=item.user_id,
                    schedule_date=item.schedule_date,
                    action="clear",
                    shift_code=record.shift_code,
                    start_time=record.start_time,
                    end_time=record.end_time,
                    reason=record.note,
                    was_locked=bool(lock and lock.is_locked),
                    changed_by=current_user.id,
                ))
                db.delete(record)
                record_map.pop((item.user_id, item.schedule_date), None)
                cleared += 1
            continue
        if lock and lock.is_locked:
            if not item.override_locked:
                raise HTTPException(status_code=409, detail="所选员工包含固定班次，请确认临时调整并填写原因")
            if not (item.note or '').strip():
                raise HTTPException(status_code=400, detail="临时调整固定班次必须填写原因")
        start_value, end_value = _storage_times(item.shift_code, item.start_time, item.end_time)
        db.add(EmployeeShiftOverrideAudit(
            user_id=item.user_id,
            schedule_date=item.schedule_date,
            action="set",
            shift_code=item.shift_code,
            start_time=start_value,
            end_time=end_value,
            reason=item.note,
            was_locked=bool(lock and lock.is_locked),
            changed_by=current_user.id,
        ))
        if record:
            record.shift_code = item.shift_code
            record.start_time = start_value
            record.end_time = end_value
            record.note = item.note
            record.updated_by = current_user.id
            record.updated_at = datetime.utcnow()
        else:
            record = EmployeeShiftOverride(
                user_id=item.user_id,
                schedule_date=item.schedule_date,
                shift_code=item.shift_code,
                start_time=start_value,
                end_time=end_value,
                note=item.note,
                updated_by=current_user.id,
            )
            db.add(record)
            record_map[(item.user_id, item.schedule_date)] = record
        updated += 1
    db.commit()
    return {"updated": updated, "cleared": cleared}


@workbench_router.get("/employee-shifts/me")
def get_my_employee_shift(
    schedule_date: date = Query(...),
    include_department: bool = Query(True),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    users = [current_user]
    if include_department and current_user.department:
        users = (
            db.query(AppUser)
            .filter(
                AppUser.is_active == True,
                AppUser.department.in_(department_filter_values(current_user.department)),
            )
            .order_by(AppUser.full_name.asc(), AppUser.username.asc())
            .all()
        )
    resolved = _resolve_employee_shift_rows(db, users, schedule_date, schedule_date)
    me = next((item for item in resolved if item["user_id"] == str(current_user.id)), None)
    return {
        "date": schedule_date.isoformat(),
        "me": me,
        "department_members": resolved if include_department and current_user.department else [],
    }


@router.get("/staff/list")
def get_staff_list(db: Session = Depends(get_db)):
    """获取所有活跃内部员工列表（用于排班页面初始化人员模板）"""
    users = db.query(AppUser).filter(AppUser.is_active == True).all()
    result = []
    for u in users:
        result.append({
            "id": str(u.id),
            "name": u.full_name or u.username,
            "dept": normalize_department(u.department) or "",
            "fixedTasks": u.fixed_tasks or [],
        })
    return result


@router.get("/translators/list")
def get_translator_list(
    direction: Optional[str] = Query(None, description="翻译方向: zh_en / en_zh / both"),
    active_only: bool = Query(False, description="仅返回活跃译员"),
    db: Session = Depends(get_db),
):
    """获取所有译员列表（含排班属性），用于译员优先次序表初始化"""
    query = db.query(Translator)
    if direction:
        query = query.filter(
            (Translator.direction == direction) | (Translator.direction == "both") | (Translator.direction == None)
        )
    if active_only:
        query = query.filter(Translator.status == "active")
    query = query.order_by(Translator.default_priority.asc())
    translators = query.all()
    result = []
    for t in translators:
        result.append({
            "id": str(t.id),
            "name": t.translator_name,
            "type": t.translation_type or "",
            "quality": t.quality_score or "",
            "cloudRev": _format_cloud_revision(t.can_cloud_edit, t.can_revision),
            "dailyRate": _format_daily_rate(t.daily_accept_count, t.hourly_speed, t.daily_word_capacity),
            "direction": t.direction or "",
            "order": str(t.default_priority) if t.default_priority else "N/A",
            "remarks": t.schedule_remarks or "",
            "status": t.status or "standby",
            "availableTimeSlot": t.available_time_slot or "",
            "dailyAcceptCount": t.daily_accept_count,
            "hourlySpeed": t.hourly_speed,
            "dailyWordCapacity": t.daily_word_capacity,
            "canCloudEdit": t.can_cloud_edit,
            "canRevision": t.can_revision,
            "domainSkills": t.domain_skills or [],
        })
    return result


@router.get("/translators/availability-grid")
def get_translator_availability_grid(
    date_from: date = Query(...),
    date_to: date = Query(...),
    direction: Optional[str] = Query(None),
    keyword: Optional[str] = Query(None),
    availability_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    if (date_to - date_from).days > 31:
        raise HTTPException(status_code=400, detail="单次最多查询 32 天排期")
    query = db.query(Translator)
    if direction:
        query = query.filter(or_(Translator.direction == direction, Translator.direction == "both", Translator.direction.is_(None)))
    if keyword and keyword.strip():
        query = query.filter(Translator.translator_name.ilike(f"%{keyword.strip()}%"))
    translators = query.order_by(Translator.default_priority.asc().nullslast(), Translator.translator_name.asc()).all()
    translator_ids = [item.id for item in translators]
    schedule_rows = db.query(TranslatorSchedule).filter(
        TranslatorSchedule.translator_id.in_(translator_ids),
        TranslatorSchedule.schedule_date >= date_from,
        TranslatorSchedule.schedule_date <= date_to,
    ).all() if translator_ids else []
    schedules = {(item.translator_id, item.schedule_date): item for item in schedule_rows}
    dates = [date_from + timedelta(days=offset) for offset in range((date_to - date_from).days + 1)]
    result = []
    for translator in translators:
        days = []
        for schedule_day in dates:
            row = schedules.get((translator.id, schedule_day))
            days.append({
                "date": schedule_day.isoformat(),
                "availability_status": row.availability_status if row else "unconfirmed",
                "available_time_slot": row.available_time_slot if row else "",
                "remaining_capacity": row.remaining_capacity if row else None,
                "remarks": row.remarks if row else "",
                "last_confirmed_at": row.last_confirmed_at.isoformat() if row and row.last_confirmed_at else None,
                "source_type": row.source_type if row else None,
            })
        if availability_status and not any(day["availability_status"] == availability_status for day in days):
            continue
        result.append({
            "translator_id": str(translator.id),
            "name": translator.translator_name,
            "direction": translator.direction or "",
            "default_priority": translator.default_priority,
            "translation_type": translator.translation_type or "",
            "quality": translator.quality_score or "",
            "can_cloud_edit": translator.can_cloud_edit,
            "can_revision": translator.can_revision,
            "daily_accept_count": translator.daily_accept_count,
            "hourly_speed": translator.hourly_speed,
            "daily_word_capacity": translator.daily_word_capacity,
            "days": days,
        })
    return {"date_from": date_from.isoformat(), "date_to": date_to.isoformat(), "translators": result}


@router.get("/translators/{translator_id}/availability", response_model=List[TranslatorScheduleResponse])
def get_translator_availability(
    translator_id: UUID,
    date_from: date = Query(..., description="开始日期"),
    date_to: date = Query(..., description="结束日期"),
    db: Session = Depends(get_db),
):
    translator = db.query(Translator).filter(Translator.id == translator_id).first()
    if not translator:
        raise HTTPException(status_code=404, detail="译员不存在")
    if date_from > date_to:
        raise HTTPException(status_code=400, detail="开始日期不能晚于结束日期")
    return (
        db.query(TranslatorSchedule)
        .filter(
            TranslatorSchedule.translator_id == translator_id,
            TranslatorSchedule.schedule_date >= date_from,
            TranslatorSchedule.schedule_date <= date_to,
        )
        .order_by(TranslatorSchedule.schedule_date.asc())
        .all()
    )


@router.put("/translators/{translator_id}/availability/{schedule_date}", response_model=TranslatorScheduleResponse)
def upsert_translator_availability(
    translator_id: UUID,
    schedule_date: date,
    data: TranslatorScheduleUpdate,
    db: Session = Depends(get_db),
):
    translator = db.query(Translator).filter(Translator.id == translator_id).first()
    if not translator:
        raise HTTPException(status_code=404, detail="译员不存在")

    record = (
        db.query(TranslatorSchedule)
        .filter(
            TranslatorSchedule.translator_id == translator_id,
            TranslatorSchedule.schedule_date == schedule_date,
        )
        .first()
    )
    payload = data.model_dump(exclude_unset=True)

    if record:
        for key, value in payload.items():
            setattr(record, key, value)
        record.updated_at = datetime.utcnow()
    else:
        record = TranslatorSchedule(
            translator_id=translator_id,
            schedule_date=schedule_date,
            **payload,
        )
        db.add(record)

    db.commit()
    db.refresh(record)
    return record


@router.post("/translators/{translator_id}/availability", response_model=TranslatorScheduleResponse, status_code=status.HTTP_201_CREATED)
def create_translator_availability(
    translator_id: UUID,
    data: TranslatorScheduleCreate,
    db: Session = Depends(get_db),
):
    translator = db.query(Translator).filter(Translator.id == translator_id).first()
    if not translator:
        raise HTTPException(status_code=404, detail="译员不存在")
    if data.translator_id != translator_id:
        raise HTTPException(status_code=400, detail="路径中的译员ID与请求体不一致")

    existing = (
        db.query(TranslatorSchedule)
        .filter(
            TranslatorSchedule.translator_id == translator_id,
            TranslatorSchedule.schedule_date == data.schedule_date,
        )
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="该日期排期已存在，请改用更新接口")

    record = TranslatorSchedule(**data.model_dump())
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.delete("/translators/{translator_id}/availability/{schedule_date}", status_code=status.HTTP_204_NO_CONTENT)
def delete_translator_availability(
    translator_id: UUID,
    schedule_date: date,
    db: Session = Depends(get_db),
):
    record = (
        db.query(TranslatorSchedule)
        .filter(
            TranslatorSchedule.translator_id == translator_id,
            TranslatorSchedule.schedule_date == schedule_date,
        )
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="该日期排期不存在")
    db.delete(record)
    db.commit()


@router.post("/translators/import-demo")
async def import_translator_schedule_demo(
    file: UploadFile = File(...),
    overwrite: bool = Form(True),
    db: Session = Depends(get_db),
):
    filename = file.filename or "translator_schedule_demo.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"暂时只支持 Excel xlsx 文件导入：{exc}") from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 内容不足，至少需要表头和一行数据")

    headers = [_normalize_text(cell) for cell in rows[0]]
    name_idx = next((i for i, h in enumerate(headers) if "姓名" in h), None)
    submit_time_idx = next((i for i, h in enumerate(headers) if "提交" in h and "时间" in h), None)
    fill_date_idx = next((i for i, h in enumerate(headers) if ("填写" in h and "日期" in h) or ("本周总" in h and "日期" in h)), None)
    remarks_idx = next((i for i, h in enumerate(headers) if "备注" in h), None)
    weekday_columns = {}
    for idx, header in enumerate(headers):
        for weekday, offset in WEEKDAY_HEADER_MAP.items():
            if weekday in header:
                weekday_columns[idx] = offset
                break

    if name_idx is None or not weekday_columns:
        raise HTTPException(status_code=400, detail="未识别到姓名列或星期列，请先保持当前导出模板格式")

    translators = db.query(Translator).all()
    translator_map = {}
    for translator in translators:
        translator_map.setdefault(_normalize_person_name(translator.translator_name), translator)

    created_count = 0
    updated_count = 0
    imported_rows = 0
    matched_translators = set()
    unmatched_names = []

    for row in rows[1:]:
        if not row or all(cell in (None, "") for cell in row):
            continue
        raw_name = row[name_idx] if name_idx < len(row) else None
        normalized_name = _normalize_person_name(raw_name)
        if not normalized_name:
            continue

        translator = translator_map.get(normalized_name)
        if not translator:
            unmatched_names.append(_normalize_text(raw_name))
            continue

        base_date = None
        if fill_date_idx is not None and fill_date_idx < len(row):
            base_date = _parse_excel_date(row[fill_date_idx])
        if base_date is None and submit_time_idx is not None and submit_time_idx < len(row):
            submit_dt = _parse_excel_datetime(row[submit_time_idx])
            if submit_dt:
                base_date = submit_dt.date()
        if base_date is None:
            continue

        week_monday = base_date - timedelta(days=base_date.weekday())
        remarks = _normalize_text(row[remarks_idx]) if remarks_idx is not None and remarks_idx < len(row) else ""
        submitted_at = None
        if submit_time_idx is not None and submit_time_idx < len(row):
            submitted_at = _parse_excel_datetime(row[submit_time_idx])

        row_imported = False
        for col_idx, weekday_offset in weekday_columns.items():
            if col_idx >= len(row):
                continue
            slot_text = _cell_to_slot_text(row[col_idx])
            if not slot_text:
                continue

            schedule_day = week_monday + timedelta(days=weekday_offset)
            record = (
                db.query(TranslatorSchedule)
                .filter(
                    TranslatorSchedule.translator_id == translator.id,
                    TranslatorSchedule.schedule_date == schedule_day,
                )
                .first()
            )
            payload = {
                "availability_status": "available",
                "available_time_slot": slot_text,
                "source_type": "excel_demo",
                "source_ref": filename,
                "last_confirmed_at": submitted_at,
                "remarks": remarks or None,
            }
            if record:
                if overwrite:
                    for key, value in payload.items():
                        setattr(record, key, value)
                    record.updated_at = datetime.utcnow()
                    updated_count += 1
                    row_imported = True
            else:
                db.add(TranslatorSchedule(
                    translator_id=translator.id,
                    schedule_date=schedule_day,
                    **payload,
                ))
                created_count += 1
                row_imported = True

        if row_imported:
            matched_translators.add(translator.translator_name)
            imported_rows += 1

    db.commit()
    return {
        "file_name": filename,
        "sheet_name": sheet.title,
        "imported_rows": imported_rows,
        "matched_translators": len(matched_translators),
        "created_records": created_count,
        "updated_records": updated_count,
        "unmatched_names": unmatched_names,
    }


@router.post("/translators/import-demo/preview")
async def preview_translator_schedule_demo(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or "translator_schedule_demo.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    parsed = _parse_translator_schedule_demo_rows(content, filename, db)
    return {
        "file_name": parsed["file_name"],
        "sheet_name": parsed["sheet_name"],
        "headers": parsed["headers"],
        "matched_translators": parsed["matched_translators"],
        "unmatched_names": parsed["unmatched_names"],
        "preview_count": len(parsed["preview_items"]),
        "preview_items": parsed["preview_items"][:200],
    }


@router.post("/translators/import-demo-v2/preview")
async def preview_translator_schedule_demo_v2(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or "translator_schedule_demo.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    parsed = _parse_translator_schedule_demo_rows_v2(content, filename, db)
    return {
        "file_name": parsed["file_name"],
        "sheet_name": parsed["sheet_name"],
        "headers": parsed["headers"],
        "matched_translators": parsed["matched_translators"],
        "unmatched_names": parsed["unmatched_names"],
        "preview_count": len(parsed["preview_items"]),
        "preview_items": parsed["preview_items"][:200],
    }


@router.post("/translators/import-demo-v2")
async def import_translator_schedule_demo_v2(
    file: UploadFile = File(...),
    overwrite: bool = Form(True),
    db: Session = Depends(get_db),
):
    filename = file.filename or "translator_schedule_demo.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    parsed = _parse_translator_schedule_demo_rows_v2(content, filename, db)
    parsed.setdefault("errors", [])
    return _apply_translator_schedule_import(parsed, overwrite, db)


def _apply_translator_schedule_import(parsed: dict, overwrite: bool, db: Session):
    created_count = 0
    updated_count = 0
    skipped_count = 0

    for item in parsed["preview_items"]:
        schedule_day = date.fromisoformat(item["schedule_date"])
        payload = {
            "availability_status": item.get("availability_status") or "available",
            "available_time_slot": item["available_time_slot"],
            "remaining_capacity": item.get("remaining_capacity"),
            "source_type": item["source_type"],
            "source_ref": item["source_ref"],
            "last_confirmed_at": datetime.fromisoformat(item["last_confirmed_at"]) if item["last_confirmed_at"] else None,
            "remarks": item["remarks"],
        }
        record = (
            db.query(TranslatorSchedule)
            .filter(
                TranslatorSchedule.translator_id == item["translator_id"],
                TranslatorSchedule.schedule_date == schedule_day,
            )
            .first()
        )
        if record:
            if overwrite:
                for key, value in payload.items():
                    setattr(record, key, value)
                record.updated_at = datetime.utcnow()
                updated_count += 1
            else:
                skipped_count += 1
        else:
            db.add(TranslatorSchedule(
                translator_id=item["translator_id"],
                schedule_date=schedule_day,
                **payload,
            ))
            created_count += 1

    db.commit()
    return {
        "file_name": parsed["file_name"],
        "sheet_name": parsed["sheet_name"],
        "imported_rows": len({item["row_no"] for item in parsed["preview_items"]}),
        "matched_translators": parsed["matched_translators"],
        "created_records": created_count,
        "updated_records": updated_count,
        "skipped_records": skipped_count,
        "unmatched_names": parsed["unmatched_names"],
        "errors": parsed.get("errors", []),
    }


@router.post("/translators/import/preview")
async def preview_translator_schedule(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    filename = file.filename or "translator_schedule.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    parsed = _parse_translator_schedule_auto(content, filename, db)
    return {
        "file_name": parsed["file_name"],
        "sheet_name": parsed["sheet_name"],
        "format": parsed.get("format"),
        "headers": parsed["headers"],
        "matched_translators": parsed["matched_translators"],
        "unmatched_names": parsed["unmatched_names"],
        "errors": parsed.get("errors", []),
        "preview_count": len(parsed["preview_items"]),
        "preview_items": parsed["preview_items"][:500],
    }


@router.post("/translators/import")
async def import_translator_schedule(
    file: UploadFile = File(...),
    overwrite: bool = Form(True),
    db: Session = Depends(get_db),
):
    filename = file.filename or "translator_schedule.xlsx"
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")
    parsed = _parse_translator_schedule_auto(content, filename, db)
    return _apply_translator_schedule_import(parsed, overwrite, db)


@router.get("/translators/import-template")
def download_translator_schedule_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "译员排期"
    sheet.append(["译员ID", "译员姓名", "日期", "接稿状态", "可接时段", "剩余容量", "备注"])
    sheet.append(["", "示例译员", date.today().isoformat(), "可接稿", "09:00-18:00", 3000, "示例行，导入前请删除"])
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="translator_schedule_template.xlsx"'},
    )


@router.get("/{schedule_date}", response_model=WorkScheduleResponse)
def get_schedule(schedule_date: date, db: Session = Depends(get_db)):
    """获取某日的排班数据"""
    record = db.query(WorkSchedule).filter(WorkSchedule.schedule_date == schedule_date).first()
    if not record:
        raise HTTPException(status_code=404, detail="该日期暂无排班数据")
    return _normalize_schedule_record(record)


@router.put("/{schedule_date}", response_model=WorkScheduleResponse)
def upsert_schedule(
    schedule_date: date,
    data: WorkScheduleUpdate,
    db: Session = Depends(get_db),
):
    """创建或更新某日的排班数据（Upsert）"""
    record = db.query(WorkSchedule).filter(WorkSchedule.schedule_date == schedule_date).first()
    payload = _normalize_schedule_payload(data.model_dump(exclude_unset=True))

    if record:
        update_data = payload
        for key, value in update_data.items():
            setattr(record, key, value)
        record.updated_at = datetime.utcnow()
    else:
        create_data = payload
        record = WorkSchedule(schedule_date=schedule_date, **create_data)
        db.add(record)

    db.commit()
    db.refresh(record)
    return _normalize_schedule_record(record)


@router.post("/copy", response_model=WorkScheduleResponse)
def copy_schedule(
    source_date: date = Query(..., description="源日期"),
    target_date: date = Query(..., description="目标日期"),
    db: Session = Depends(get_db),
):
    """将某一天的排班复制到另外一天（用于"从昨日复制"功能）"""
    source = db.query(WorkSchedule).filter(WorkSchedule.schedule_date == source_date).first()
    if not source:
        raise HTTPException(status_code=404, detail="源日期暂无排班数据")

    normalized_dept_person_data = [_normalize_dept_person(p) for p in (source.dept_person_data or [])]
    normalized_not_scheduled_tasks = [_normalize_not_scheduled(i) for i in (source.not_scheduled_tasks or [])]

    existing = db.query(WorkSchedule).filter(WorkSchedule.schedule_date == target_date).first()
    if existing:
        existing.leave_notes = source.leave_notes
        existing.dept_person_data = normalized_dept_person_data
        existing.not_scheduled_tasks = normalized_not_scheduled_tasks
        existing.pm_rotation_order = source.pm_rotation_order
        existing.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(existing)
        return _normalize_schedule_record(existing)
    else:
        new_record = WorkSchedule(
            schedule_date=target_date,
            leave_notes=source.leave_notes,
            dept_person_data=normalized_dept_person_data,
            not_scheduled_tasks=normalized_not_scheduled_tasks,
            pm_rotation_order=source.pm_rotation_order,
        )
        db.add(new_record)
        db.commit()
        db.refresh(new_record)
        return _normalize_schedule_record(new_record)


@router.delete("/{schedule_date}", status_code=status.HTTP_204_NO_CONTENT)
def delete_schedule(schedule_date: date, db: Session = Depends(get_db)):
    """删除某日的排班数据"""
    record = db.query(WorkSchedule).filter(WorkSchedule.schedule_date == schedule_date).first()
    if not record:
        raise HTTPException(status_code=404, detail="该日期暂无排班数据")
    db.delete(record)
    db.commit()
